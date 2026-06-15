"""
Smart Finance Core - Financial Reports API
업로드된 계정별 원장 데이터 기반 재무보고서 (기간 기반)

- 더존 6자리 계정코드 지원 (000101 → 101 → 자산)
- 여러 파일을 나눠 업로드해도 같은 기간이면 합산
- 데이터 소스: CONFIRMED Voucher 단일 소스 (ai_raw는 스테이징 전용)
"""
import json
import logging
import math
import re
from datetime import datetime, date as _date
from calendar import monthrange
from decimal import Decimal
from io import BytesIO
from typing import Optional, List
from urllib.parse import quote

from fastapi import APIRouter, Depends, HTTPException, Query
from pydantic import BaseModel
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, and_, or_, update
from starlette.responses import StreamingResponse

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

from app.core.database import get_db
from app.core.security import get_current_user
from app.models.user import User
from app.models.ai import AIRawTransactionData, AIDataUploadHistory
from app.models.accounting import Account, AccountCategory, Voucher, VoucherLine, VoucherStatus
from app.services.unified_ledger import (
    unified_aggregation_subquery,
    unified_rows_subquery,
    has_closing_vouchers,
)
from app.core.account_category import (
    categorize_account,
    CATEGORY_LABEL,
    strip_code,
    is_bs_account,
)

router = APIRouter()


# ============ 더존 계정코드 분류 ============

# 더존 표준 계정과목 코드 체계 (앞자리 0 제거 후)
# 1xx: 자산 (10x~14x 유동, 15x~19x 비유동)
# 2xx: 부채 (20x~26x 유동, 27x~29x 비유동)
# 3xx: 자본
# 4xx: 매출/수익
# 5xx: 매출원가
# 8xx: 판매비와관리비
# 9xx: 영업외손익

DOUZONE_CATEGORY = {
    '1': '자산', '2': '부채', '3': '자본',
    '4': '수익', '5': '매출원가',
    '6': '제조원가', '7': '제조원가',
    '8': '판관비', '9': '영업외',
}


def _strip_code(code: str) -> str:
    """더존 6자리 코드에서 앞의 0 제거: '000101' → '101'"""
    return strip_code(code)


def _classify_code(code: str) -> tuple:
    """
    더존 계정코드 분류.
    Returns: (first_digit, second_digit, stripped_code, category_name)
    """
    s = strip_code(code)
    first = s[0] if s else '0'
    second = int(s[1]) if len(s) > 1 and s[1].isdigit() else 0
    cat = DOUZONE_CATEGORY.get(first, '미분류')
    return first, second, s, cat


# ============ Helpers ============

def _period_bounds(year: Optional[int], month: Optional[int]):
    """year/month → (period_start, period_end) date 객체 반환. 둘 다 None이면 (None, None)."""
    if year and month:
        return (_date(year, month, 1), _date(year, month, monthrange(year, month)[1]))
    elif year:
        return (_date(year, 1, 1), _date(year, 12, 31))
    return (None, None)


async def _resolve_names(db: AsyncSession, codes: list, mode: str = "multi") -> dict:
    """계정 코드 → 이름 매핑. Voucher/Account 테이블 기반."""
    if not codes:
        return {}

    names = {}

    # 1) unified(voucher) subquery에서 source_account_name 조회
    sub = unified_aggregation_subquery()
    raw_result = await db.execute(
        select(
            sub.c.source_account_code,
            func.max(sub.c.source_account_name).label("name"),
        )
        .where(
            sub.c.source_account_code.in_(codes),
            sub.c.source_account_name.isnot(None),
            sub.c.source_account_name != "",
        )
        .group_by(sub.c.source_account_code)
    )
    names.update({r.source_account_code: r.name for r in raw_result.all()})

    # 2) Account 테이블에서 보충
    missing = [c for c in codes if c not in names]
    if missing:
        result = await db.execute(
            select(Account.code, Account.name).where(Account.code.in_(missing))
        )
        names.update({r.code: r.name for r in result.all()})

    # 3) 최종 fallback
    for c in codes:
        if c not in names or not names[c]:
            names[c] = f"계정 {strip_code(c)}"
    return names


async def _get_account_balances(
    db: AsyncSession,
    year: Optional[int] = None,
    month: Optional[int] = None,
    exclude_closing: bool = False,
) -> list:
    """
    계정별 잔액 집계 — CONFIRMED Voucher 단일 소스.
    year/month 지정 시 그 기간만, 둘 다 None이면 전체.
    """
    period_start, period_end = _period_bounds(year, month)

    sub = unified_aggregation_subquery(period_start, period_end, exclude_closing=exclude_closing)
    q = select(
        sub.c.source_account_code.label("code"),
        func.coalesce(func.sum(sub.c.debit_amount), 0).label("debit_total"),
        func.coalesce(func.sum(sub.c.credit_amount), 0).label("credit_total"),
        func.count().label("tx_count"),
    ).group_by(sub.c.source_account_code).order_by(sub.c.source_account_code)

    result = await db.execute(q)
    return result.all()


async def _detect_years(db: AsyncSession) -> List[int]:
    """DB에 있는 모든 연도 감지 (CONFIRMED Voucher transaction_date 기준)"""
    years = set()

    # 1) Voucher.transaction_date에서 연도 추출
    result = await db.execute(
        select(func.distinct(func.extract('year', Voucher.transaction_date)))
        .where(
            Voucher.status == VoucherStatus.CONFIRMED,
            Voucher.transaction_date.isnot(None),
        )
    )
    for row in result.all():
        if row[0] is not None:
            years.add(int(row[0]))

    # 2) AIDataUploadHistory.created_at에서 연도 추출 (fallback — staging uploads)
    from app.models.ai import UploadStatus
    upload_years_result = await db.execute(
        select(func.distinct(func.extract('year', AIDataUploadHistory.created_at)))
        .where(
            AIDataUploadHistory.status == UploadStatus.COMPLETED,
            AIDataUploadHistory.created_at.isnot(None),
        )
    )
    for row in upload_years_result.all():
        if row[0] is not None:
            years.add(int(row[0]))

    return sorted(years, reverse=True)


# ============ Endpoints ============

@router.get("/available-years")
async def get_available_years(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """조회 가능한 연도 목록"""
    years = await _detect_years(db)

    from app.models.ai import UploadStatus
    uploads_result = await db.execute(
        select(AIDataUploadHistory)
        .where(AIDataUploadHistory.status == UploadStatus.COMPLETED)
        .order_by(AIDataUploadHistory.created_at.desc())
        .limit(50)
    )
    uploads = uploads_result.scalars().all()

    # ai_raw 행 수 — 스테이징 전용 통계
    total_rows = await db.scalar(
        select(func.count(AIRawTransactionData.id))
    ) or 0

    return {
        "years": years,
        "total_raw_rows": total_rows,
        "uploads": [
            {
                "id": u.id,
                "filename": u.filename,
                "saved_count": u.saved_count or 0,
                "upload_type": u.upload_type,
                "created_at": u.created_at.isoformat() if u.created_at else None,
            }
            for u in uploads
        ],
    }


@router.get("/summary")
async def get_financial_summary(
    year: Optional[int] = Query(None),
    upload_id: Optional[int] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """재무 요약 (기간 기반) — Voucher 단일 소스"""
    period_start, period_end = _period_bounds(year, None)
    sub = unified_aggregation_subquery(period_start, period_end)

    q = select(
        func.coalesce(func.sum(sub.c.debit_amount), 0).label("total_debit"),
        func.coalesce(func.sum(sub.c.credit_amount), 0).label("total_credit"),
        func.count().label("tx_count"),
    )

    totals = await db.execute(q)
    t = totals.one()

    years = await _detect_years(db)

    return {
        "year": year,
        "total_debit": float(t.total_debit),
        "total_credit": float(t.total_credit),
        "total_transactions": t.tx_count,
        "ledger_mode": "multi",
        "available_years": years,
    }


@router.get("/trial-balance")
async def get_trial_balance(
    year: Optional[int] = Query(None),
    month: Optional[int] = Query(None, ge=1, le=12, description="월 필터 (1-12)"),
    upload_id: Optional[int] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """시산표 - 계정별 차변/대변 합계 (기간 기반, 더존 코드 분류)"""
    # 시산표는 마감전표 포함(기본) — BS 잔액과 일치시키기 위해
    rows = await _get_account_balances(db, year=year, month=month, exclude_closing=False)

    codes = [r.code for r in rows]
    names = await _resolve_names(db, codes)

    items = []
    total_debit = 0.0
    total_credit = 0.0
    for r in rows:
        d = float(r.debit_total)
        c = float(r.credit_total)
        total_debit += d
        total_credit += c
        code = r.code
        first, _, stripped, cat_name = _classify_code(code)
        items.append({
            "account_code": code,
            "account_name": names.get(code, f"계정 {stripped}"),
            "category_name": cat_name,
            "debit_total": d,
            "credit_total": c,
            "balance": d - c,
            "tx_count": r.tx_count,
        })

    return {
        "year": year,
        "month": month,
        "items": items,
        "total_debit": total_debit,
        "total_credit": total_credit,
        "ledger_mode": "multi",
    }


@router.get("/income-statement")
async def get_income_statement(
    year: Optional[int] = Query(None),
    month: Optional[int] = Query(None),
    upload_id: Optional[int] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """손익계산서 - 더존 계정코드 기준 (기간 기반). 마감전표 이중반영 차단."""
    if year is None:
        years = await _detect_years(db)
        if years:
            year = years[0]

    # 손익 보고서: 마감전표(auto_closing) 제외 — 이중반영 방지
    period_start, period_end = _period_bounds(year, month)
    _has_closing = await has_closing_vouchers(db, period_start, period_end)
    rows = await _get_account_balances(db, year=year, month=month, exclude_closing=True)

    codes = [r.code for r in rows]
    names = await _resolve_names(db, codes)

    revenue_items = []
    cogs_items = []
    sga_items = []
    non_op_income_items = []
    non_op_expense_items = []

    # 위하고 손익계산서 로직: 45x(제품매출원가/상품매출원가)에 결산 분개가 있으면
    # 5xx/6xx/7xx/8xx(제)는 이미 45x로 흡수된 원가요소 → 매출원가에서 제외 (이중계산 방지).
    # 45x가 비어있고 마감전표도 없으면(결산 전, 진행중인 월) 5xx/6xx/7xx/8xx(제) raw 합 사용.
    has_45x_cogs = False
    for r in rows:
        code = r.code
        if code and len(code) >= 3 and code[0] == '4':
            try:
                ci = int(code[:3])
                if 450 <= ci < 500 and float(r.debit_total) > 0:
                    has_45x_cogs = True
                    break
            except ValueError:
                pass

    # 마감전표 존재 시: 5xx~8xx(제) 원가요소는 이미 대체분개로 처리됨 → skip
    skip_cost_elements = has_45x_cogs or _has_closing

    for r in rows:
        code = r.code
        d = float(r.debit_total)
        c = float(r.credit_total)
        first, _, stripped, _ = _classify_code(code)
        acct_name = names.get(code, f"계정 {stripped}")

        item = {"code": code, "name": acct_name, "debit": d, "credit": c, "tx_count": r.tx_count}

        # 4xx 안에서 매출 vs 매출원가 구분 (위하고: 45x = 제품/상품매출원가)
        is_4xx_cogs = False
        if first == '4' and code and len(code) >= 3:
            try:
                ci = int(code[:3])
                if 450 <= ci < 500:
                    is_4xx_cogs = True
            except ValueError:
                pass
        # 8xx 중 이름에 "(제)" 포함 = 제조원가 요소 (801 급여(제) 등)
        is_manuf_8xx = first == '8' and "(제)" in (acct_name or "")
        # 5/6/7xx + 8xx(제) 는 원가요소 → 45x 결산 완료 또는 마감전표 존재 시 제외
        is_cost_element = first in ('5', '6', '7') or is_manuf_8xx

        # voucher 단일 소스 → 항상 multi 모드 계산 (source_account_code 기준)
        if first == '4':
            if is_4xx_cogs:
                item["amount"] = d - c
                cogs_items.append(item)
            else:
                item["amount"] = c - d
                revenue_items.append(item)
        elif is_cost_element:
            if not skip_cost_elements:  # 결산 전 = 원가요소 직접 합산
                item["amount"] = d - c
                cogs_items.append(item)
            # else: 45x 결산 완료 또는 마감전표 존재 — 이중계산 방지로 skip
        elif first == '8':  # (판) 등 일반 판관비
            item["amount"] = d - c
            sga_items.append(item)
        elif first == '9':
            net = c - d
            if net >= 0:
                item["amount"] = net
                non_op_income_items.append(item)
            else:
                item["amount"] = -net
                non_op_expense_items.append(item)

    for arr in [revenue_items, cogs_items, sga_items, non_op_income_items, non_op_expense_items]:
        arr.sort(key=lambda x: x["amount"], reverse=True)

    revenue_total = sum(i["amount"] for i in revenue_items)
    cogs_total = sum(i["amount"] for i in cogs_items)
    gross_profit = revenue_total - cogs_total
    sga_total = sum(i["amount"] for i in sga_items)
    operating_income = gross_profit - sga_total
    non_op_income_total = sum(i["amount"] for i in non_op_income_items)
    non_op_expense_total = sum(i["amount"] for i in non_op_expense_items)
    pre_tax_income = operating_income + non_op_income_total - non_op_expense_total
    tax = 0
    net_income = pre_tax_income - tax

    def pct(val):
        return round(val / revenue_total * 100, 2) if revenue_total else 0.0

    def make_items(arr):
        return [{"code": i["code"], "name": i["name"], "amount": i["amount"],
                 "debit": i.get("debit", 0), "credit": i.get("credit", 0),
                 "tx_count": i.get("tx_count", 0)} for i in arr]

    sections = [
        {"id": "I", "name": "매출액", "items": make_items(revenue_items), "total": revenue_total, "pct": 100.0 if revenue_total > 0 else 0.0},
        {"id": "II", "name": "매출원가", "items": make_items(cogs_items), "total": cogs_total, "pct": pct(cogs_total)},
        {"id": "III", "name": "매출총이익", "items": [], "total": gross_profit, "pct": pct(gross_profit), "is_subtotal": True},
        {"id": "IV", "name": "판매비와관리비", "items": make_items(sga_items), "total": sga_total, "pct": pct(sga_total)},
        {"id": "V", "name": "영업이익", "items": [], "total": operating_income, "pct": pct(operating_income), "is_subtotal": True},
        {"id": "VI", "name": "영업외수익", "items": make_items(non_op_income_items), "total": non_op_income_total, "pct": pct(non_op_income_total)},
        {"id": "VII", "name": "영업외비용", "items": make_items(non_op_expense_items), "total": non_op_expense_total, "pct": pct(non_op_expense_total)},
        {"id": "VIII", "name": "법인세차감전순이익", "items": [], "total": pre_tax_income, "pct": pct(pre_tax_income), "is_subtotal": True},
        {"id": "IX", "name": "법인세등", "items": [], "total": tax, "pct": pct(tax)},
        {"id": "X", "name": "당기순이익", "items": [], "total": net_income, "pct": pct(net_income), "is_subtotal": True},
    ]

    return {
        "year": year,
        "month": month,
        "ledger_mode": "multi",
        "sections": sections,
        "revenue_total": revenue_total,
        "net_income": net_income,
    }


@router.get("/balance-sheet")
async def get_balance_sheet(
    year: Optional[int] = Query(None),
    upload_id: Optional[int] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """재무상태표 - 더존 계정코드 기준 (기간 기반). BS는 exclude_closing=False."""
    rows = await _get_account_balances(db, year=year, month=None, exclude_closing=False)

    codes = [r.code for r in rows]
    names = await _resolve_names(db, codes)

    current_asset_items = []
    noncurrent_asset_items = []
    current_liab_items = []
    noncurrent_liab_items = []
    equity_items = []

    for r in rows:
        code = r.code
        d = float(r.debit_total)
        c = float(r.credit_total)
        first, second, stripped, _ = _classify_code(code)
        acct_name = names.get(code, f"계정 {stripped}")

        if first == '1':
            amount = d - c  # 자산: 차변 잔액
            item = {"code": code, "name": acct_name, "amount": amount}
            if second <= 4:
                current_asset_items.append(item)
            else:
                noncurrent_asset_items.append(item)
        elif first == '2':
            amount = c - d  # 부채: 대변 잔액
            item = {"code": code, "name": acct_name, "amount": amount}
            if second <= 6:
                current_liab_items.append(item)
            else:
                noncurrent_liab_items.append(item)
        elif first == '3':
            amount = c - d  # 자본: 대변 잔액
            equity_items.append({"code": code, "name": acct_name, "amount": amount})

    current_asset_total = sum(i["amount"] for i in current_asset_items)
    noncurrent_asset_total = sum(i["amount"] for i in noncurrent_asset_items)
    total_assets = current_asset_total + noncurrent_asset_total
    current_liab_total = sum(i["amount"] for i in current_liab_items)
    noncurrent_liab_total = sum(i["amount"] for i in noncurrent_liab_items)
    total_liabilities = current_liab_total + noncurrent_liab_total
    equity_total = sum(i["amount"] for i in equity_items)

    sections = [
        {"id": "assets", "name": "자산", "subsections": [
            {"name": "I. 유동자산", "items": current_asset_items, "total": current_asset_total},
            {"name": "II. 비유동자산", "items": noncurrent_asset_items, "total": noncurrent_asset_total},
        ], "total": total_assets},
        {"id": "liabilities", "name": "부채", "subsections": [
            {"name": "I. 유동부채", "items": current_liab_items, "total": current_liab_total},
            {"name": "II. 비유동부채", "items": noncurrent_liab_items, "total": noncurrent_liab_total},
        ], "total": total_liabilities},
        {"id": "equity", "name": "자본", "subsections": [
            {"name": "자본 항목", "items": equity_items, "total": equity_total},
        ], "total": equity_total},
    ]

    return {
        "year": year,
        "ledger_mode": "multi",
        "sections": sections,
        "total_assets": total_assets,
        "total_liabilities": total_liabilities,
        "total_equity": equity_total,
    }


@router.get("/balance-sheet-monthly")
async def get_balance_sheet_monthly(
    year: int = Query(..., ge=2020, le=2030),
    upload_id: Optional[int] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    월별 재무상태표 — 해당 연도 1~12월 각 월말의 누적 BS 시계열.
    한 번의 SQL로 (월, 계정) 그룹화 후 코드 단 누적으로 효율적 처리.
    데이터 없는 월은 skip — 마지막 데이터 월까지만 반환.
    BS는 마감전표 포함(exclude_closing=False).
    """
    from collections import defaultdict

    period_start = _date(year, 1, 1)
    period_end = _date(year, 12, 31)

    # unified subquery — 연도 전체, BS는 마감전표 포함
    sub = unified_aggregation_subquery(period_start, period_end, exclude_closing=False)

    # 월 키: transaction_date는 'YYYY-MM-DD' 문자열 → substr(1,7) = 'YYYY-MM'
    month_key = func.substr(sub.c.transaction_date, 1, 7)

    # (월, 계정) 합산
    monthly_rows = (await db.execute(
        select(
            month_key.label('ym'),
            sub.c.source_account_code.label('code'),
            func.coalesce(func.sum(sub.c.debit_amount), 0).label('debit'),
            func.coalesce(func.sum(sub.c.credit_amount), 0).label('credit'),
        )
        .where(sub.c.source_account_code.isnot(None), sub.c.source_account_code != '')
        .group_by(month_key, sub.c.source_account_code)
    )).all()

    if not monthly_rows:
        return {"year": year, "ledger_mode": "multi", "months": []}

    # 월별 그룹화
    monthly_data: dict = defaultdict(dict)  # {ym: {code: {debit, credit}}}
    all_codes: set = set()
    months_with_data: set = set()
    for r in monthly_rows:
        if not r.ym or not r.code:
            continue
        monthly_data[r.ym][r.code] = {'debit': float(r.debit or 0), 'credit': float(r.credit or 0)}
        all_codes.add(r.code)
        months_with_data.add(r.ym)

    # 자산·부채·자본 계정만 BS 대상 (1xx, 2xx, 3xx)
    bs_codes = [c for c in all_codes if _classify_code(c)[0] in ('1', '2', '3')]
    names = await _resolve_names(db, bs_codes)

    # 마지막 데이터 월
    last_m = max(int(ym.split('-')[1]) for ym in months_with_data)

    months_result = []
    cumulative: dict = {code: {'debit': 0.0, 'credit': 0.0} for code in bs_codes}

    for m in range(1, last_m + 1):
        ym = f"{year}-{m:02d}"
        # 이번 월 데이터 누적
        month_data = monthly_data.get(ym, {})
        for code in bs_codes:
            if code in month_data:
                cumulative[code]['debit'] += month_data[code]['debit']
                cumulative[code]['credit'] += month_data[code]['credit']

        # BS 분류
        current_asset_items, noncurrent_asset_items = [], []
        current_liab_items, noncurrent_liab_items = [], []
        equity_items = []

        for code in bs_codes:
            d = cumulative[code]['debit']
            c = cumulative[code]['credit']
            if d == 0 and c == 0:
                continue
            first, second, stripped, _cat = _classify_code(code)
            acct_name = names.get(code, f"계정 {stripped}")

            if first == '1':
                amount = d - c  # 자산: 차변 잔액 (voucher=multi 모드)
                item = {"code": code, "name": acct_name, "amount": amount}
                if second <= 4:
                    current_asset_items.append(item)
                else:
                    noncurrent_asset_items.append(item)
            elif first == '2':
                amount = c - d  # 부채: 대변 잔액
                item = {"code": code, "name": acct_name, "amount": amount}
                if second <= 6:
                    current_liab_items.append(item)
                else:
                    noncurrent_liab_items.append(item)
            elif first == '3':
                amount = c - d  # 자본: 대변 잔액
                equity_items.append({"code": code, "name": acct_name, "amount": amount})

        for items in (current_asset_items, noncurrent_asset_items,
                       current_liab_items, noncurrent_liab_items, equity_items):
            items.sort(key=lambda x: -abs(x["amount"]))

        ca_total = sum(i["amount"] for i in current_asset_items)
        nca_total = sum(i["amount"] for i in noncurrent_asset_items)
        cl_total = sum(i["amount"] for i in current_liab_items)
        ncl_total = sum(i["amount"] for i in noncurrent_liab_items)
        eq_total = sum(i["amount"] for i in equity_items)

        last_day = monthrange(year, m)[1]

        months_result.append({
            "month": m,
            "month_label": f"{year}-{m:02d}",
            "month_end": f"{year}-{m:02d}-{last_day:02d}",
            "sections": [
                {"id": "assets", "name": "자산", "subsections": [
                    {"name": "I. 유동자산", "items": current_asset_items, "total": ca_total},
                    {"name": "II. 비유동자산", "items": noncurrent_asset_items, "total": nca_total},
                ], "total": ca_total + nca_total},
                {"id": "liabilities", "name": "부채", "subsections": [
                    {"name": "I. 유동부채", "items": current_liab_items, "total": cl_total},
                    {"name": "II. 비유동부채", "items": noncurrent_liab_items, "total": ncl_total},
                ], "total": cl_total + ncl_total},
                {"id": "equity", "name": "자본", "subsections": [
                    {"name": "자본 항목", "items": equity_items, "total": eq_total},
                ], "total": eq_total},
            ],
            "total_assets": ca_total + nca_total,
            "total_liabilities": cl_total + ncl_total,
            "total_equity": eq_total,
        })

    return {
        "year": year,
        "ledger_mode": "multi",
        "months": months_result,
    }


@router.get("/monthly-trend")
async def get_monthly_trend(
    year: Optional[int] = Query(None),
    account_code: Optional[str] = Query(None),
    upload_id: Optional[int] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """월별 추이 (기간 기반) — Voucher 단일 소스"""
    period_start, period_end = _period_bounds(year, None)
    sub = unified_aggregation_subquery(period_start, period_end)

    # account_code 필터
    extra_where = []
    if account_code:
        extra_where.append(sub.c.source_account_code == account_code)

    # transaction_date는 'YYYY-MM-DD' 형식 → substr(1,7) = 'YYYY-MM'
    month_key = func.substr(sub.c.transaction_date, 1, 7)

    q = select(
        month_key.label('ym'),
        func.coalesce(func.sum(sub.c.debit_amount), 0).label('debit'),
        func.coalesce(func.sum(sub.c.credit_amount), 0).label('credit'),
        func.count().label('cnt'),
    )
    if extra_where:
        q = q.where(*extra_where)
    q = q.group_by(month_key).order_by(month_key)

    result = await db.execute(q)
    rows = result.all()

    data = []
    for r in rows:
        if not r.ym:
            continue
        debit = float(r.debit or 0)
        credit = float(r.credit or 0)
        data.append({
            "month": r.ym,
            "debit_total": debit,
            "credit_total": credit,
            "net": debit - credit,
            "tx_count": r.cnt,
        })

    return {
        "year": year,
        "account_code": account_code,
        "data": data,
    }


@router.get("/account-detail")
async def get_account_detail(
    account_code: str = Query(...),
    year: Optional[int] = Query(None),
    month: Optional[int] = Query(None, ge=1, le=12, description="월 필터 (1-12)"),
    upload_id: Optional[int] = Query(None),
    page: int = Query(default=1, ge=1),
    size: int = Query(default=50, ge=1, le=500),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """계정별 거래 상세 조회 (기간 기반) — Voucher 단일 소스"""
    period_start, period_end = _period_bounds(year, month)
    sub = unified_rows_subquery(period_start, period_end)

    base_where = [sub.c.source_account_code == account_code]

    total = await db.scalar(
        select(func.count()).select_from(sub).where(*base_where)
    ) or 0
    total_pages = max(1, math.ceil(total / size))

    summary_result = await db.execute(
        select(
            func.coalesce(func.sum(sub.c.debit_amount), 0).label("debit_total"),
            func.coalesce(func.sum(sub.c.credit_amount), 0).label("credit_total"),
        ).select_from(sub).where(*base_where)
    )
    s = summary_result.one()

    offset = (page - 1) * size
    data_result = await db.execute(
        select(
            sub.c.transaction_date,
            sub.c.description,
            sub.c.merchant_name,
            sub.c.debit_amount,
            sub.c.credit_amount,
            sub.c.source_account_code,
            sub.c.row_number,
        )
        .select_from(sub)
        .where(*base_where)
        .order_by(sub.c.transaction_date, sub.c.row_number)
        .offset(offset)
        .limit(size)
    )
    rows = data_result.all()

    return {
        "account_code": account_code,
        "year": year,
        "month": month,
        "total": total,
        "page": page,
        "size": size,
        "total_pages": total_pages,
        "items": [
            {
                "row_number": r.row_number,
                "transaction_date": r.transaction_date,
                "description": r.description,
                "merchant_name": r.merchant_name,
                "debit_amount": float(r.debit_amount or 0),
                "credit_amount": float(r.credit_amount or 0),
                "account_code": None,          # voucher는 상대계정 NULL (unified_rows 스펙)
                "source_account_code": r.source_account_code,
            }
            for r in rows
        ],
        "summary": {
            "debit_total": float(s.debit_total),
            "credit_total": float(s.credit_total),
            "balance": float(s.debit_total) - float(s.credit_total),
        },
    }


@router.get("/account-monthly")
async def get_account_monthly_breakdown(
    account_code: str = Query(..., description="계정코드"),
    year: Optional[int] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """특정 계정의 월별 집계 (시산표 드릴다운용) — Voucher 단일 소스"""
    period_start, period_end = _period_bounds(year, None)
    sub = unified_aggregation_subquery(period_start, period_end)

    month_key = func.substr(sub.c.transaction_date, 1, 7)

    result = await db.execute(
        select(
            month_key.label('ym'),
            func.coalesce(func.sum(sub.c.debit_amount), 0).label('debit'),
            func.coalesce(func.sum(sub.c.credit_amount), 0).label('credit'),
            func.count().label('cnt'),
        )
        .where(sub.c.source_account_code == account_code)
        .group_by(month_key)
        .order_by(month_key)
    )
    rows = result.all()

    months_data = []
    for r in rows:
        if not r.ym:
            continue
        try:
            m_num = int(r.ym.split("-")[1])
        except (IndexError, ValueError):
            continue
        debit = float(r.debit or 0)
        credit = float(r.credit or 0)
        months_data.append({
            "month": m_num,
            "month_label": f"{m_num}월",
            "debit_total": debit,
            "credit_total": credit,
            "balance": debit - credit,
            "tx_count": r.cnt,
        })

    total_debit = sum(m["debit_total"] for m in months_data)
    total_credit = sum(m["credit_total"] for m in months_data)
    total_count = sum(m["tx_count"] for m in months_data)

    return {
        "account_code": account_code,
        "year": year,
        "months": months_data,
        "total": {
            "debit_total": total_debit,
            "credit_total": total_credit,
            "balance": total_debit - total_credit,
            "tx_count": total_count,
        },
    }


@router.get("/account-detail/export/excel")
async def export_account_detail_excel(
    account_code: str = Query(...),
    year: Optional[int] = Query(None),
    month: Optional[int] = Query(None, ge=1, le=12),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """계정별 분개 내역 엑셀 다운로드 — Voucher 단일 소스"""
    period_start, period_end = _period_bounds(year, month)
    sub = unified_rows_subquery(period_start, period_end)

    data_result = await db.execute(
        select(
            sub.c.transaction_date,
            sub.c.description,
            sub.c.merchant_name,
            sub.c.debit_amount,
            sub.c.credit_amount,
            sub.c.source_account_code,
            sub.c.row_number,
        )
        .select_from(sub)
        .where(sub.c.source_account_code == account_code)
        .order_by(sub.c.transaction_date, sub.c.row_number)
    )
    rows = data_result.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "분개내역"

    header = ["날짜", "적요", "거래처", "차변", "대변", "계정코드"]
    ws.append(header)

    for col_idx, _ in enumerate(header, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    total_debit = 0.0
    total_credit = 0.0

    for r in rows:
        debit = float(r.debit_amount or 0)
        credit = float(r.credit_amount or 0)
        total_debit += debit
        total_credit += credit
        ws.append([
            r.transaction_date or "",
            r.description or "",
            r.merchant_name or "",
            debit if debit else None,
            credit if credit else None,
            r.source_account_code or "",
        ])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=4, max_col=5):
        for cell in row:
            if cell.value is not None:
                cell.number_format = "#,##0"

    summary_row = ws.max_row + 1
    ws.cell(row=summary_row, column=1, value="합계")
    ws.cell(row=summary_row, column=4, value=total_debit)
    ws.cell(row=summary_row, column=5, value=total_credit)
    for col_idx in [1, 4, 5]:
        cell = ws.cell(row=summary_row, column=col_idx)
        cell.font = Font(bold=True)
    for col_idx in [4, 5]:
        ws.cell(row=summary_row, column=col_idx).number_format = "#,##0"

    col_widths = [15, 40, 20, 18, 18, 12]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"분개내역_{account_code}_{year or 'all'}{'_' + str(month) + '월' if month else ''}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )


# ============ 계정명 보정 (기존 데이터용) ============

class AccountNameMapping(BaseModel):
    code: str
    name: str

class BackfillNamesRequest(BaseModel):
    mappings: List[AccountNameMapping]

@router.post("/backfill-names")
async def backfill_account_names(
    data: BackfillNamesRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    기존 업로드 데이터의 source_account_name 보정.
    프론트에서 엑셀 파일의 [CODE] NAME 매핑만 추출해서 전송.
    (ai_raw는 스테이징 전용이지만 이름 보정은 여기서 유지)
    """
    updated = 0
    for m in data.mappings:
        result = await db.execute(
            update(AIRawTransactionData)
            .where(
                AIRawTransactionData.source_account_code == m.code,
                or_(
                    AIRawTransactionData.source_account_name.is_(None),
                    AIRawTransactionData.source_account_name == "",
                )
            )
            .values(source_account_name=m.name)
        )
        updated += result.rowcount

    await db.commit()
    return {"updated_rows": updated, "codes_processed": len(data.mappings)}


@router.get("/debug-data")
async def debug_raw_data(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """DB에 저장된 원본 데이터 구조 확인용 (ai_raw 스테이징 + Voucher 통계)"""
    # ai_raw 총 행 수 (스테이징 전용)
    total = await db.scalar(select(func.count(AIRawTransactionData.id))) or 0

    # source_account_code 분포 (ai_raw staging)
    src_codes = await db.execute(
        select(
            AIRawTransactionData.source_account_code,
            func.count(AIRawTransactionData.id).label("cnt"),
        )
        .group_by(AIRawTransactionData.source_account_code)
        .order_by(func.count(AIRawTransactionData.id).desc())
        .limit(20)
    )
    source_code_dist = [{"code": r.source_account_code, "count": r.cnt} for r in src_codes.all()]

    # account_code 분포 (ai_raw staging)
    acct_codes = await db.execute(
        select(
            AIRawTransactionData.account_code,
            func.count(AIRawTransactionData.id).label("cnt"),
        )
        .group_by(AIRawTransactionData.account_code)
        .order_by(func.count(AIRawTransactionData.id).desc())
        .limit(20)
    )
    account_code_dist = [{"code": r.account_code, "count": r.cnt} for r in acct_codes.all()]

    # 샘플 행 5개 (ai_raw staging)
    sample = await db.execute(
        select(AIRawTransactionData).order_by(AIRawTransactionData.id).limit(5)
    )
    sample_rows = [
        {
            "id": r.id,
            "source_account_code": r.source_account_code,
            "source_account_name": getattr(r, 'source_account_name', None),
            "account_code": r.account_code,
            "account_name": r.account_name,
            "original_description": r.original_description[:80] if r.original_description else None,
            "debit_amount": float(r.debit_amount),
            "credit_amount": float(r.credit_amount),
            "transaction_date": r.transaction_date,
        }
        for r in sample.scalars().all()
    ]

    # Voucher 통계
    voucher_total = await db.scalar(
        select(func.count(Voucher.id)).where(Voucher.status == VoucherStatus.CONFIRMED)
    ) or 0

    distinct_src = await db.scalar(
        select(func.count(func.distinct(Account.code)))
        .select_from(VoucherLine)
        .join(Account, VoucherLine.account_id == Account.id)
        .join(Voucher, VoucherLine.voucher_id == Voucher.id)
        .where(Voucher.status == VoucherStatus.CONFIRMED)
    ) or 0

    return {
        "total_rows": total,
        "detected_mode": "multi",
        "distinct_source_account_codes": distinct_src,
        "distinct_account_codes": 0,
        "top_source_codes": source_code_dist,
        "top_account_codes": account_code_dist,
        "sample_rows": sample_rows,
        "voucher_confirmed_count": voucher_total,
    }


# ============ AI 재무 분석 ============

@router.get("/ai-analysis")
async def get_ai_analysis(
    year: Optional[int] = Query(None),
    month: Optional[int] = Query(None, ge=1, le=12),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """AI 재무 분석 - 4대 카테고리 (Claude Opus 4.6)"""
    from app.core.config import settings

    if not settings.ANTHROPIC_API_KEY:
        raise HTTPException(
            status_code=400,
            detail="AI 분석을 위해 ANTHROPIC_API_KEY 환경변수를 설정해주세요. (Railway Variables에서 설정)"
        )

    # 1) 재무 데이터 수집
    if year is None:
        years = await _detect_years(db)
        if years:
            year = years[0]

    period_label = f"{year}년 {month}월" if month else f"{year}년"
    # 손익용: 마감전표 제외
    period_start, period_end = _period_bounds(year, month)
    _has_closing = await has_closing_vouchers(db, period_start, period_end)
    rows = await _get_account_balances(db, year=year, month=month, exclude_closing=True)
    codes = [r.code for r in rows]
    names = await _resolve_names(db, codes)

    # 2) 손익계산서 데이터 구성
    revenue_total = 0.0
    cogs_total = 0.0
    sga_total = 0.0
    non_op_income_total = 0.0
    non_op_expense_total = 0.0
    revenue_items_text = []
    cogs_items_text = []
    sga_items_text = []
    non_op_items_text = []

    has_45x_cogs = False
    for r in rows:
        code = r.code
        if code and len(code) >= 3 and code[0] == '4':
            try:
                ci = int(code[:3])
                if 450 <= ci < 500 and float(r.debit_total) > 0:
                    has_45x_cogs = True
                    break
            except ValueError:
                pass

    skip_cost_elements = has_45x_cogs or _has_closing

    for r in rows:
        code = r.code
        d = float(r.debit_total)
        c = float(r.credit_total)
        first, _, stripped, _ = _classify_code(code)
        acct_name = names.get(code, f"계정 {stripped}")

        is_4xx_cogs = False
        if first == '4' and code and len(code) >= 3:
            try:
                ci = int(code[:3])
                if 450 <= ci < 500:
                    is_4xx_cogs = True
            except ValueError:
                pass
        is_manuf_8xx = first == '8' and "(제)" in (acct_name or "")
        is_cost_element = first in ('5', '6', '7') or is_manuf_8xx

        if first == '4':
            if is_4xx_cogs:
                amt = d - c
                cogs_total += amt
                cogs_items_text.append(f"  {acct_name}({code}): {amt:,.0f}")
            else:
                amt = c - d
                revenue_total += amt
                revenue_items_text.append(f"  {acct_name}({code}): {amt:,.0f}")
        elif is_cost_element:
            if not skip_cost_elements:
                amt = d - c
                cogs_total += amt
                cogs_items_text.append(f"  {acct_name}({code}): {amt:,.0f}")
        elif first == '8':
            amt = d - c
            sga_total += amt
            sga_items_text.append(f"  {acct_name}({code}): {amt:,.0f}")
        elif first == '9':
            net = c - d
            if net >= 0:
                non_op_income_total += net
            else:
                non_op_expense_total += (-net)
            non_op_items_text.append(f"  {acct_name}({code}): 차변={d:,.0f} 대변={c:,.0f}")

    gross_profit = revenue_total - cogs_total
    operating_income = gross_profit - sga_total
    net_income = operating_income + non_op_income_total - non_op_expense_total

    # 3) 재무상태표 데이터 (BS는 마감전표 포함)
    bs_rows = await _get_account_balances(db, year=year, month=month, exclude_closing=False)
    bs_names = await _resolve_names(db, [r.code for r in bs_rows])

    asset_items = []
    liab_items = []
    equity_items = []
    total_assets = 0.0
    total_liab = 0.0
    total_equity = 0.0

    for r in bs_rows:
        code = r.code
        d = float(r.debit_total)
        c = float(r.credit_total)
        first, _, stripped, _ = _classify_code(code)
        acct_name = bs_names.get(code, f"계정 {stripped}")

        if first == '1':
            amt = d - c
            total_assets += amt
            if abs(amt) > 10_000_000:
                asset_items.append(f"  {acct_name}({code}): {amt:,.0f}")
        elif first == '2':
            amt = c - d
            total_liab += amt
            if abs(amt) > 10_000_000:
                liab_items.append(f"  {acct_name}({code}): {amt:,.0f}")
        elif first == '3':
            amt = c - d
            total_equity += amt
            equity_items.append(f"  {acct_name}({code}): {amt:,.0f}")

    # 4) 시산표 요약 (미분류 등 특이 항목)
    unclassified = []
    large_balance = []
    for r in rows:
        code = r.code
        d = float(r.debit_total)
        c = float(r.credit_total)
        first, _, stripped, _ = _classify_code(code)
        acct_name = names.get(code, f"계정 {stripped}")
        balance = d - c

        if DOUZONE_CATEGORY.get(first) is None:
            unclassified.append(f"  {acct_name}({code}): 차변={d:,.0f} 대변={c:,.0f}")
        if abs(balance) > 500_000_000:
            large_balance.append(f"  {acct_name}({code}): 잔액={balance:,.0f}")

    # 5) 프롬프트 구성
    prompt = f"""당신은 한국 중소기업 전문 공인회계사입니다. 아래 {period_label} 재무 데이터를 분석해주세요.
회사명: 주식회사 조인앤조인 (식품/유통업)
분석 기간: {period_label}

=== 손익계산서 ===
매출액: {revenue_total:,.0f}원
{chr(10).join(revenue_items_text[:20])}

매출원가: {cogs_total:,.0f}원
{chr(10).join(cogs_items_text[:20])}

매출총이익: {gross_profit:,.0f}원
매출총이익률: {(gross_profit/revenue_total*100) if revenue_total else 0:.1f}%

판매비와관리비: {sga_total:,.0f}원
{chr(10).join(sga_items_text[:20])}

영업이익: {operating_income:,.0f}원
영업이익률: {(operating_income/revenue_total*100) if revenue_total else 0:.1f}%

영업외수익: {non_op_income_total:,.0f}원
영업외비용: {non_op_expense_total:,.0f}원
당기순이익: {net_income:,.0f}원

=== 재무상태표 ===
자산 총계: {total_assets:,.0f}원
주요 자산:
{chr(10).join(asset_items[:15])}

부채 총계: {total_liab:,.0f}원
주요 부채:
{chr(10).join(liab_items[:15])}

자본 총계: {total_equity:,.0f}원
{chr(10).join(equity_items[:10])}

부채비율: {(total_liab/total_equity*100) if total_equity else 0:.1f}%

=== 특이사항 ===
미분류 계정: {len(unclassified)}개
{chr(10).join(unclassified[:10])}

잔액 5억 이상 계정: {len(large_balance)}개
{chr(10).join(large_balance[:15])}

총 계정 수: {len(rows)}개, 거래 건수: {sum(r.tx_count for r in rows):,}건

---
위 데이터를 아래 4가지 카테고리로 분석해주세요. 각 카테고리별 3~5개의 구체적 항목을 제시하세요.

반드시 아래 JSON 형식으로만 응답하세요 (다른 텍스트 없이):
{{
  "financial_improvements": [
    {{"title": "항목명", "current": "현재 수치/상황", "issue": "문제점", "recommendation": "구체적 개선방안"}}
  ],
  "pl_improvements": [
    {{"title": "항목명", "current": "현재 수치/상황", "issue": "문제점", "recommendation": "구체적 개선방안"}}
  ],
  "account_notable": [
    {{"title": "항목명", "current": "현재 수치/상황", "issue": "특이사항 설명", "recommendation": "확인/조치 방안"}}
  ],
  "accounting_concerns": [
    {{"title": "항목명", "current": "현재 수치/상황", "issue": "우려 사항", "recommendation": "확인/조치 방안"}}
  ]
}}"""

    # 6) Claude API 호출
    try:
        import anthropic
        client = anthropic.Anthropic(api_key=settings.ANTHROPIC_API_KEY)
        message = client.messages.create(
            model=settings.ANTHROPIC_MODEL,
            max_tokens=8000,
            temperature=0.3,
            system="당신은 한국 중소기업 전문 공인회계사입니다. 반드시 요청된 JSON 형식으로만 응답하세요. JSON 외 다른 텍스트는 절대 포함하지 마세요.",
            messages=[
                {"role": "user", "content": prompt},
            ],
        )
        content = ""
        for block in (message.content or []):
            if getattr(block, 'type', None) == 'text':
                content = block.text
                break
        if not content:
            content = message.content[0].text if message.content else "{}"

        import re as _re
        parsed = None

        try:
            parsed = json.loads(content.strip())
        except json.JSONDecodeError:
            pass

        if parsed is None:
            json_match = _re.search(r'```(?:json)?\s*([\s\S]*?)```', content)
            if json_match:
                try:
                    parsed = json.loads(json_match.group(1).strip())
                except json.JSONDecodeError:
                    pass

        if parsed is None:
            brace_match = _re.search(r'\{[\s\S]*\}', content)
            if brace_match:
                try:
                    parsed = json.loads(brace_match.group(0))
                except json.JSONDecodeError:
                    pass

        if parsed is None:
            logger.error(f"AI 응답 JSON 파싱 실패: {content[:1000]}")
            raise HTTPException(status_code=500, detail="AI 응답 파싱 실패")

        analysis = parsed
    except Exception as e:
        logger.error(f"AI 분석 오류: {e}")
        raise HTTPException(status_code=500, detail=f"AI 분석 오류: {str(e)}")

    return {
        "year": year,
        "month": month,
        "period": period_label,
        "analysis": analysis,
        "generated_at": datetime.now().isoformat(),
        "summary": {
            "revenue": revenue_total,
            "net_income": net_income,
            "total_assets": total_assets,
            "total_liabilities": total_liab,
            "account_count": len(rows),
        },
    }


# ============ AI 계정 분개 점검 ============

@router.get("/ai-account-check")
async def ai_account_check(
    year: Optional[int] = Query(None),
    account_codes: Optional[str] = Query(None, description="Comma-separated account codes"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """AI 계정 분개 점검 - 선택된 계정코드의 분개 내역을 분석하여 올바른 분류인지 검증 — Voucher 단일 소스"""
    from app.core.config import settings

    if not settings.ANTHROPIC_API_KEY:
        raise HTTPException(
            status_code=400,
            detail="AI 분석을 위해 ANTHROPIC_API_KEY 환경변수를 설정해주세요. (Railway Variables에서 설정)"
        )

    if not account_codes:
        raise HTTPException(
            status_code=400,
            detail="검토할 계정코드를 지정해주세요. (account_codes 파라미터)"
        )

    code_list = [c.strip() for c in account_codes.split(",") if c.strip()]
    if not code_list:
        raise HTTPException(
            status_code=400,
            detail="유효한 계정코드가 없습니다."
        )

    period_start, period_end = _period_bounds(year, None)
    period_label = f"{year}년" if year else "전체 기간"

    account_sections = []
    for code in code_list:
        sub = unified_rows_subquery(period_start, period_end)

        result = await db.execute(
            select(
                sub.c.description,
                sub.c.merchant_name,
                sub.c.debit_amount,
                sub.c.credit_amount,
                sub.c.transaction_date,
                sub.c.source_account_code,
                sub.c.source_account_name,
            )
            .select_from(sub)
            .where(sub.c.source_account_code == code)
            .order_by(sub.c.transaction_date)
            .limit(50)
        )
        txns = result.all()

        if not txns:
            continue

        name_map = await _resolve_names(db, [code])
        acct_name = name_map.get(code, f"계정 {strip_code(code)}")

        tx_lines = []
        for tx in txns:
            date_str = tx.transaction_date or "날짜없음"
            desc = tx.description or ""
            merchant = tx.merchant_name or ""
            debit = float(tx.debit_amount or 0)
            credit = float(tx.credit_amount or 0)

            tx_lines.append(
                f"  - {date_str} | {desc} | {merchant} | "
                f"차변: {debit:,.0f} / 대변: {credit:,.0f}"
            )

        account_sections.append({
            "code": code,
            "name": acct_name,
            "tx_count": len(txns),
            "text": f"계정코드: {code} / 계정명: {acct_name}\n거래 내역:\n" + "\n".join(tx_lines),
        })

    if not account_sections:
        raise HTTPException(
            status_code=404,
            detail="지정된 계정코드에 대한 거래 데이터가 없습니다."
        )

    accounts_text = "\n\n".join(section["text"] for section in account_sections)

    prompt = f"""당신은 한국 중소기업 전문 공인회계사입니다. 아래 선택된 계정의 분개 내역을 검토해주세요.
회사명: 주식회사 조인앤조인 (식품/유통업)
분석 기간: {period_label}

=== 검토 대상 계정 ===
{accounts_text}

---

각 계정별로 아래 항목을 점검해주세요:

[A. 분개 정확성 점검]
1. 계정 분류가 올바른지 (해당 거래가 이 계정에 맞는지)
2. 상대 계정이 적절한지
3. 중복 분개가 있는지
4. 금액이 비정상적인 거래가 있는지

[B. 내부통제 및 부정 징후 점검 — 매우 중요]
5. 자금 유출 의심: 정당한 사유 없이 외부로 자금이 빠져나가는 패턴 (가공 거래처, 비정상 지급, 수수료 과다 등)
6. 잘못된 이체/계좌 오류: 실제 거래와 다른 계좌로 기록되거나, 동일 금액이 다른 계정으로 분산 기록된 경우
7. 담당자 의도적 실수 가능성: 소액 분산 처리하여 결재 한도를 회피하는 패턴, 기말/기초에 몰린 비정상 거래, 동일 거래처에 반복적 소액 지급, 적요가 모호하거나 일반적이지 않은 표현, 주말/공휴일 처리 거래
8. 회전거래(라운드트리핑) 의심: 동일 또는 유사 금액이 입출금 반복되는 패턴
9. 유령 거래처: 특정 거래처에 대한 거래가 비정상적으로 집중되거나, 거래처명이 불명확한 경우
10. 기타 내부통제 취약점

반드시 아래 JSON 형식으로만 응답하세요 (다른 텍스트 없이):
{{
  "accounts": [
    {{
      "code": "계정코드",
      "name": "계정명",
      "status": "정상" | "확인필요" | "문제발견",
      "findings": [
        {{"type": "misclassification|duplicate|unusual_amount|wrong_counterpart|fund_leakage|wrong_transfer|intentional_error|round_tripping|ghost_vendor|internal_control|other", "description": "구체적 설명", "severity": "high|medium|low", "transaction_detail": "관련 거래 날짜와 금액 포함", "recommendation": "권장 조치 (경영진 보고 필요 여부 포함)"}}
      ],
      "summary": "해당 계정에 대한 종합 의견 (부정 위험도 포함)"
    }}
  ],
  "overall_summary": "전체적인 분개 점검 결과 요약 (내부통제 관점 포함)"
}}"""

    try:
        import anthropic
        client = anthropic.Anthropic(api_key=settings.ANTHROPIC_API_KEY)
        message = client.messages.create(
            model=settings.ANTHROPIC_MODEL,
            max_tokens=8000,
            temperature=0.3,
            system="당신은 한국 중소기업 전문 공인회계사이자 내부감사/포렌식 회계 전문가입니다. 분개의 정확성뿐 아니라 자금 유출, 횡령, 담당자 부정행위 징후를 날카롭게 포착해야 합니다. 의심스러운 패턴이 있으면 반드시 지적하세요. 반드시 요청된 JSON 형식으로만 응답하세요. JSON 외 다른 텍스트는 절대 포함하지 마세요.",
            messages=[
                {"role": "user", "content": prompt},
            ],
        )

        content = ""
        for block in (message.content or []):
            if getattr(block, 'type', None) == 'text':
                content = block.text
                break
        if not content:
            content = message.content[0].text if message.content else "{}"

        import re as _re
        parsed = None

        try:
            parsed = json.loads(content.strip())
        except json.JSONDecodeError:
            pass

        if parsed is None:
            json_match = _re.search(r'```(?:json)?\s*([\s\S]*?)```', content)
            if json_match:
                try:
                    parsed = json.loads(json_match.group(1).strip())
                except json.JSONDecodeError:
                    pass

        if parsed is None:
            brace_match = _re.search(r'\{[\s\S]*\}', content)
            if brace_match:
                try:
                    parsed = json.loads(brace_match.group(0))
                except json.JSONDecodeError:
                    pass

        if parsed is None:
            logger.error(f"AI 계정점검 응답 JSON 파싱 실패: {content[:1000]}")
            raise HTTPException(status_code=500, detail="AI 응답 파싱 실패")

        analysis = parsed
    except Exception as e:
        logger.error(f"AI 계정 분개 점검 오류: {e}")
        raise HTTPException(status_code=500, detail=f"AI 계정 분개 점검 오류: {str(e)}")

    return {
        "year": year,
        "period": period_label,
        "account_codes": code_list,
        "analysis": analysis,
        "generated_at": datetime.now().isoformat(),
        "accounts_checked": len(account_sections),
        "total_transactions_checked": sum(s["tx_count"] for s in account_sections),
    }
