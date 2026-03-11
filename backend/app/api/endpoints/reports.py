"""
Smart Finance Core - Reports API
보고서 및 엑셀 내보내기 API 엔드포인트
"""
from datetime import date, datetime
from typing import Optional
from io import BytesIO
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db

router = APIRouter()


@router.get("/vouchers/excel")
async def export_vouchers_excel(
    from_date: date,
    to_date: date,
    department_id: Optional[int] = None,
    status: Optional[str] = None,
    db: AsyncSession = Depends(get_db)
):
    """
    전표 목록 엑셀 내보내기

    - 기간/부서/상태별 전표를 엑셀 파일로 다운로드합니다
    - 더존 양식과 호환됩니다
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side
    from app.models.accounting import Voucher, VoucherStatus
    from sqlalchemy import select, and_

    # 데이터 조회
    conditions = [
        Voucher.voucher_date >= from_date,
        Voucher.voucher_date <= to_date
    ]
    if department_id:
        conditions.append(Voucher.department_id == department_id)
    if status:
        conditions.append(Voucher.status == VoucherStatus(status))

    result = await db.execute(
        select(Voucher).where(and_(*conditions)).order_by(Voucher.voucher_date)
    )
    vouchers = result.scalars().all()

    # 엑셀 생성
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "전표목록"

    # 헤더
    headers = ["전표번호", "전표일자", "거래일자", "적요", "차변", "대변", "상태", "부서"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # 데이터
    for row, voucher in enumerate(vouchers, 2):
        ws.cell(row=row, column=1, value=voucher.voucher_number)
        ws.cell(row=row, column=2, value=voucher.voucher_date.isoformat())
        ws.cell(row=row, column=3, value=voucher.transaction_date.isoformat())
        ws.cell(row=row, column=4, value=voucher.description)
        ws.cell(row=row, column=5, value=float(voucher.total_debit))
        ws.cell(row=row, column=6, value=float(voucher.total_credit))
        ws.cell(row=row, column=7, value=voucher.status.value)
        ws.cell(row=row, column=8, value=voucher.department_id)

    # 열 너비 조정
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 2

    # 파일 저장
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"vouchers_{from_date}_{to_date}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/budget-vs-actual/excel")
async def export_budget_vs_actual_excel(
    fiscal_year: int,
    department_id: Optional[int] = None,
    db: AsyncSession = Depends(get_db)
):
    """예산 대 실적 엑셀 내보내기"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from app.services.budget_service import BudgetService

    service = BudgetService(db)
    data = await service.get_budget_vs_actual(fiscal_year, department_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "예산vs실적"

    # 헤더
    headers = ["부서", "계정코드", "계정명", "예산", "실적", "차이", "차이율(%)"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # 데이터
    for row, item in enumerate(data["items"], 2):
        ws.cell(row=row, column=1, value=item["department"])
        ws.cell(row=row, column=2, value=item["account_code"])
        ws.cell(row=row, column=3, value=item["account_name"])
        ws.cell(row=row, column=4, value=float(item["budget"]))
        ws.cell(row=row, column=5, value=float(item["actual"]))
        ws.cell(row=row, column=6, value=float(item["variance"]))
        ws.cell(row=row, column=7, value=float(item["variance_pct"]))

    # 합계
    totals = data["totals"]
    last_row = len(data["items"]) + 2
    ws.cell(row=last_row, column=3, value="합계").font = Font(bold=True)
    ws.cell(row=last_row, column=4, value=float(totals["budget"]))
    ws.cell(row=last_row, column=5, value=float(totals["actual"]))
    ws.cell(row=last_row, column=6, value=float(totals["variance"]))
    ws.cell(row=last_row, column=7, value=float(totals["variance_pct"]))

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"budget_vs_actual_{fiscal_year}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/aging/excel")
async def export_aging_excel(
    report_type: str = Query(..., pattern="^(receivables|payables)$"),
    as_of_date: Optional[date] = None,
    db: AsyncSession = Depends(get_db)
):
    """연령 분석 엑셀 내보내기"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from app.services.treasury_manager import TreasuryManager

    manager = TreasuryManager(db)

    if report_type == "receivables":
        data = await manager.get_ar_aging_report(as_of_date)
        title = "매출채권 연령분석"
        party_label = "거래처"
    else:
        data = await manager.get_ap_aging_report(as_of_date)
        title = "매입채무 연령분석"
        party_label = "공급업체"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title

    # 헤더
    headers = [party_label, "만기전", "1-30일", "31-60일", "61-90일", "90일초과", "합계"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

    # 데이터
    for row, item in enumerate(data["items"], 2):
        name = item.get("customer") or item.get("vendor", "")
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=float(item["current"]))
        ws.cell(row=row, column=3, value=float(item["days_1_30"]))
        ws.cell(row=row, column=4, value=float(item["days_31_60"]))
        ws.cell(row=row, column=5, value=float(item["days_61_90"]))
        ws.cell(row=row, column=6, value=float(item["days_over_90"]))
        ws.cell(row=row, column=7, value=float(item["total"]))

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    report_date = as_of_date or date.today()
    filename = f"aging_{report_type}_{report_date}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.post("/douzone-export")
async def export_to_douzone_format(
    from_date: date,
    to_date: date,
    export_type: str = Query("excel", pattern="^(excel|api)$"),
    db: AsyncSession = Depends(get_db)
):
    """
    더존 양식으로 내보내기

    - 더존 Smart A / iCube 업로드용 양식으로 변환합니다
    """
    import openpyxl
    from app.models.accounting import Voucher, VoucherLine, VoucherStatus
    from sqlalchemy import select, and_
    from sqlalchemy.orm import selectinload

    # 확정된 전표만 (lines와 account를 eager loading)
    result = await db.execute(
        select(Voucher)
        .options(selectinload(Voucher.lines).selectinload(VoucherLine.account))
        .where(
            and_(
                Voucher.voucher_date >= from_date,
                Voucher.voucher_date <= to_date,
                Voucher.status == VoucherStatus.CONFIRMED
            )
        ).order_by(Voucher.voucher_date)
    )
    vouchers = result.scalars().unique().all()

    # 더존 양식 엑셀 생성
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "일반전표"

    # 더존 양식 헤더
    headers = [
        "전표일자", "전표번호", "순번", "차대구분",
        "계정코드", "계정명", "적요", "금액",
        "거래처코드", "거래처명", "부서코드"
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    row = 2
    for voucher in vouchers:
        for line in voucher.lines:
            # 차변
            if line.debit_amount > 0:
                ws.cell(row=row, column=1, value=voucher.voucher_date.strftime("%Y%m%d"))
                ws.cell(row=row, column=2, value=voucher.voucher_number)
                ws.cell(row=row, column=3, value=line.line_number)
                ws.cell(row=row, column=4, value="1")  # 1=차변
                ws.cell(row=row, column=5, value=line.account.code if line.account else "")
                ws.cell(row=row, column=6, value=line.account.name if line.account else "")
                ws.cell(row=row, column=7, value=line.description or voucher.description)
                ws.cell(row=row, column=8, value=float(line.debit_amount))
                ws.cell(row=row, column=9, value=line.counterparty_business_number or "")
                ws.cell(row=row, column=10, value=line.counterparty_name or "")
                ws.cell(row=row, column=11, value=line.cost_center_code or "")
                row += 1

            # 대변
            if line.credit_amount > 0:
                ws.cell(row=row, column=1, value=voucher.voucher_date.strftime("%Y%m%d"))
                ws.cell(row=row, column=2, value=voucher.voucher_number)
                ws.cell(row=row, column=3, value=line.line_number)
                ws.cell(row=row, column=4, value="2")  # 2=대변
                ws.cell(row=row, column=5, value=line.account.code if line.account else "")
                ws.cell(row=row, column=6, value=line.account.name if line.account else "")
                ws.cell(row=row, column=7, value=line.description or voucher.description)
                ws.cell(row=row, column=8, value=float(line.credit_amount))
                ws.cell(row=row, column=9, value=line.counterparty_business_number or "")
                ws.cell(row=row, column=10, value=line.counterparty_name or "")
                ws.cell(row=row, column=11, value=line.cost_center_code or "")
                row += 1

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"douzone_export_{from_date}_{to_date}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
