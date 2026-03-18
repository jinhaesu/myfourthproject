"""
Smart Finance Core - Survey & Commute API
설문조사 & 출퇴근 관리 API 엔드포인트
"""
from datetime import datetime, date
from typing import Optional, List
import json
import io

from fastapi import APIRouter, Depends, HTTPException, status, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, delete, func, and_, or_, Integer, cast
from sqlalchemy.orm import selectinload

from app.core.database import get_db
from app.core.security import get_current_user
from app.models.user import User, Department
from app.models.survey import (
    Survey, SurveyQuestion, SurveyResponse as SurveyResponseModel,
    SurveyAnswer, CommuteRecord,
    SurveyCategory, SurveyStatus, QuestionType, TransportMethod,
)
from app.schemas.survey import (
    SurveyCreate, SurveyUpdate, SurveyResponse, SurveyListResponse,
    SurveyQuestionCreate, SurveyQuestionResponse,
    SurveyResponseCreate, SurveyResponseUpdate, SurveyResponseRecord,
    SurveyResponseListItem, SurveyAnswerCreate, SurveyAnswerResponse,
    CommuteRecordCreate, CommuteRecordUpdate, CommuteRecordResponse,
    CommuteSummaryItem, CheckInRequest, CheckOutRequest,
    SurveyTemplateInfo, TemplateCreateRequest,
)

router = APIRouter(prefix="/survey", tags=["설문/출퇴근"])


# ============================================================================
# Helper functions
# ============================================================================

def survey_to_response(survey: Survey) -> dict:
    return {
        "id": survey.id,
        "title": survey.title,
        "description": survey.description,
        "category": survey.category.value if hasattr(survey.category, "value") else str(survey.category),
        "status": survey.status.value if hasattr(survey.status, "value") else str(survey.status),
        "target_department_id": survey.target_department_id,
        "is_anonymous": survey.is_anonymous,
        "is_required": survey.is_required,
        "start_date": survey.start_date,
        "end_date": survey.end_date,
        "is_recurring": survey.is_recurring,
        "recurrence_type": survey.recurrence_type,
        "created_by": survey.created_by,
        "created_at": survey.created_at,
        "updated_at": survey.updated_at,
        "questions": [question_to_response(q) for q in (survey.questions or [])],
    }


def question_to_response(q: SurveyQuestion) -> dict:
    return {
        "id": q.id,
        "survey_id": q.survey_id,
        "question_text": q.question_text,
        "question_type": q.question_type.value if hasattr(q.question_type, "value") else str(q.question_type),
        "order": q.order,
        "is_required": q.is_required,
        "options": q.options,
        "scale_min": q.scale_min,
        "scale_max": q.scale_max,
        "scale_min_label": q.scale_min_label,
        "scale_max_label": q.scale_max_label,
        "placeholder": q.placeholder,
        "created_at": q.created_at,
    }


def response_record_to_dict(r: SurveyResponseModel, include_answers: bool = False) -> dict:
    result = {
        "id": r.id,
        "survey_id": r.survey_id,
        "user_id": r.user_id,
        "response_date": r.response_date,
        "submitted_at": r.submitted_at,
        "updated_at": r.updated_at,
    }
    if include_answers:
        result["answers"] = [answer_to_dict(a) for a in (r.answers or [])]
    return result


def answer_to_dict(a: SurveyAnswer) -> dict:
    return {
        "id": a.id,
        "response_id": a.response_id,
        "question_id": a.question_id,
        "answer_text": a.answer_text,
        "answer_number": a.answer_number,
        "answer_json": a.answer_json,
        "created_at": a.created_at,
        "updated_at": a.updated_at,
    }


def commute_to_dict(c: CommuteRecord) -> dict:
    return {
        "id": c.id,
        "user_id": c.user_id,
        "record_date": c.record_date,
        "record_year": c.record_year,
        "record_month": c.record_month,
        "check_in_time": c.check_in_time,
        "check_out_time": c.check_out_time,
        "departure_time": c.departure_time,
        "arrival_time": c.arrival_time,
        "transport_method": (
            c.transport_method.value if c.transport_method and hasattr(c.transport_method, "value")
            else str(c.transport_method) if c.transport_method else None
        ),
        "commute_duration_minutes": c.commute_duration_minutes,
        "note": c.note,
        "survey_response_id": c.survey_response_id,
        "is_late": c.is_late,
        "is_early_leave": c.is_early_leave,
        "is_absent": c.is_absent,
        "created_at": c.created_at,
        "updated_at": c.updated_at,
        "user_name": c.user.full_name if c.user else None,
        "department_name": (
            c.user.department.name
            if c.user and c.user.department else None
        ),
    }


# ============================================================================
# Survey CRUD
# ============================================================================

@router.get("/", summary="설문조사 목록 조회")
async def list_surveys(
    category: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=100),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """설문조사 목록을 조회합니다."""
    query = select(Survey).options(selectinload(Survey.questions))

    if category:
        try:
            cat_enum = SurveyCategory(category)
            query = query.where(Survey.category == cat_enum)
        except ValueError:
            pass

    if status:
        try:
            stat_enum = SurveyStatus(status)
            query = query.where(Survey.status == stat_enum)
        except ValueError:
            pass

    query = query.order_by(Survey.created_at.desc())
    offset = (page - 1) * size
    query = query.offset(offset).limit(size)

    result = await db.execute(query)
    surveys = result.scalars().all()

    # Count responses for each survey
    items = []
    for s in surveys:
        # Get response count
        count_q = select(func.count(SurveyResponseModel.id)).where(
            SurveyResponseModel.survey_id == s.id
        )
        count_result = await db.execute(count_q)
        response_count = count_result.scalar() or 0

        item = survey_to_response(s)
        item["question_count"] = len(s.questions or [])
        item["response_count"] = response_count
        items.append(item)

    return {"items": items, "page": page, "size": size, "total": len(items)}


@router.post("/", summary="설문조사 생성")
async def create_survey(
    data: SurveyCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """새 설문조사를 생성합니다."""
    try:
        category_enum = SurveyCategory(data.category)
    except ValueError:
        category_enum = SurveyCategory.GENERAL

    survey = Survey(
        title=data.title,
        description=data.description,
        category=category_enum,
        status=SurveyStatus.DRAFT,
        target_department_id=data.target_department_id,
        is_anonymous=data.is_anonymous,
        is_required=data.is_required,
        start_date=data.start_date,
        end_date=data.end_date,
        is_recurring=data.is_recurring,
        recurrence_type=data.recurrence_type,
        created_by=current_user.id,
    )
    db.add(survey)
    await db.flush()  # get survey.id

    # Add questions
    for i, q_data in enumerate(data.questions or []):
        try:
            q_type = QuestionType(q_data.question_type)
        except ValueError:
            q_type = QuestionType.TEXT

        question = SurveyQuestion(
            survey_id=survey.id,
            question_text=q_data.question_text,
            question_type=q_type,
            order=q_data.order if q_data.order is not None else i,
            is_required=q_data.is_required,
            options=q_data.options,
            scale_min=q_data.scale_min,
            scale_max=q_data.scale_max,
            scale_min_label=q_data.scale_min_label,
            scale_max_label=q_data.scale_max_label,
            placeholder=q_data.placeholder,
        )
        db.add(question)

    await db.commit()
    await db.refresh(survey)

    # Reload with relationships
    result = await db.execute(
        select(Survey)
        .options(selectinload(Survey.questions))
        .where(Survey.id == survey.id)
    )
    survey = result.scalar_one()
    return survey_to_response(survey)


@router.get("/{survey_id}", summary="설문조사 상세 조회")
async def get_survey(
    survey_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """설문조사 상세 정보를 조회합니다."""
    result = await db.execute(
        select(Survey)
        .options(selectinload(Survey.questions))
        .where(Survey.id == survey_id)
    )
    survey = result.scalar_one_or_none()
    if not survey:
        raise HTTPException(status_code=404, detail="설문조사를 찾을 수 없습니다.")
    return survey_to_response(survey)


@router.put("/{survey_id}", summary="설문조사 수정")
async def update_survey(
    survey_id: int,
    data: SurveyUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """설문조사를 수정합니다. questions를 제공하면 기존 문항을 교체합니다."""
    result = await db.execute(
        select(Survey)
        .options(selectinload(Survey.questions))
        .where(Survey.id == survey_id)
    )
    survey = result.scalar_one_or_none()
    if not survey:
        raise HTTPException(status_code=404, detail="설문조사를 찾을 수 없습니다.")

    # Update fields
    update_data = data.model_dump(exclude_unset=True, exclude={"questions"})
    for field, value in update_data.items():
        if field == "category" and value is not None:
            try:
                value = SurveyCategory(value)
            except ValueError:
                continue
        if field == "status" and value is not None:
            try:
                value = SurveyStatus(value)
            except ValueError:
                continue
        setattr(survey, field, value)

    survey.updated_at = datetime.utcnow()

    # Replace questions if provided
    if data.questions is not None:
        await db.execute(
            delete(SurveyQuestion).where(SurveyQuestion.survey_id == survey_id)
        )
        for i, q_data in enumerate(data.questions):
            try:
                q_type = QuestionType(q_data.question_type)
            except ValueError:
                q_type = QuestionType.TEXT

            question = SurveyQuestion(
                survey_id=survey_id,
                question_text=q_data.question_text,
                question_type=q_type,
                order=q_data.order if q_data.order is not None else i,
                is_required=q_data.is_required,
                options=q_data.options,
                scale_min=q_data.scale_min,
                scale_max=q_data.scale_max,
                scale_min_label=q_data.scale_min_label,
                scale_max_label=q_data.scale_max_label,
                placeholder=q_data.placeholder,
            )
            db.add(question)

    await db.commit()

    result = await db.execute(
        select(Survey)
        .options(selectinload(Survey.questions))
        .where(Survey.id == survey_id)
    )
    survey = result.scalar_one()
    return survey_to_response(survey)


@router.delete("/{survey_id}", summary="설문조사 삭제/보관")
async def delete_survey(
    survey_id: int,
    hard_delete: bool = Query(False, description="True면 완전 삭제, False면 archived 처리"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """설문조사를 삭제하거나 보관 처리합니다."""
    result = await db.execute(select(Survey).where(Survey.id == survey_id))
    survey = result.scalar_one_or_none()
    if not survey:
        raise HTTPException(status_code=404, detail="설문조사를 찾을 수 없습니다.")

    if hard_delete:
        await db.delete(survey)
    else:
        survey.status = SurveyStatus.ARCHIVED
        survey.updated_at = datetime.utcnow()

    await db.commit()
    return {"success": True, "message": "설문조사가 처리되었습니다.", "survey_id": survey_id}


@router.post("/{survey_id}/publish", summary="설문조사 게시 (active)")
async def publish_survey(
    survey_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """설문조사 상태를 active로 변경합니다."""
    result = await db.execute(select(Survey).where(Survey.id == survey_id))
    survey = result.scalar_one_or_none()
    if not survey:
        raise HTTPException(status_code=404, detail="설문조사를 찾을 수 없습니다.")

    survey.status = SurveyStatus.ACTIVE
    survey.updated_at = datetime.utcnow()
    await db.commit()
    return {"success": True, "message": "설문조사가 게시되었습니다.", "status": "active"}


@router.post("/{survey_id}/close", summary="설문조사 마감")
async def close_survey(
    survey_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """설문조사 상태를 closed로 변경합니다."""
    result = await db.execute(select(Survey).where(Survey.id == survey_id))
    survey = result.scalar_one_or_none()
    if not survey:
        raise HTTPException(status_code=404, detail="설문조사를 찾을 수 없습니다.")

    survey.status = SurveyStatus.CLOSED
    survey.updated_at = datetime.utcnow()
    await db.commit()
    return {"success": True, "message": "설문조사가 마감되었습니다.", "status": "closed"}


# ============================================================================
# Survey Responses
# ============================================================================

@router.post("/responses/", summary="설문 응답 제출")
async def submit_response(
    data: SurveyResponseCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """설문 응답을 제출합니다."""
    # Validate survey exists
    s_result = await db.execute(select(Survey).where(Survey.id == data.survey_id))
    survey = s_result.scalar_one_or_none()
    if not survey:
        raise HTTPException(status_code=404, detail="설문조사를 찾을 수 없습니다.")

    user_id = data.user_id if data.user_id else current_user.id
    resp_date = data.response_date if data.response_date else date.today()

    response = SurveyResponseModel(
        survey_id=data.survey_id,
        user_id=user_id,
        response_date=resp_date,
        submitted_at=datetime.utcnow(),
    )
    db.add(response)
    await db.flush()

    for ans_data in data.answers:
        answer = SurveyAnswer(
            response_id=response.id,
            question_id=ans_data.question_id,
            answer_text=ans_data.answer_text,
            answer_number=ans_data.answer_number,
            answer_json=ans_data.answer_json,
        )
        db.add(answer)

    await db.commit()

    result = await db.execute(
        select(SurveyResponseModel)
        .options(selectinload(SurveyResponseModel.answers))
        .where(SurveyResponseModel.id == response.id)
    )
    response = result.scalar_one()
    return response_record_to_dict(response, include_answers=True)


@router.get("/responses/", summary="설문 응답 목록 조회")
async def list_responses(
    survey_id: Optional[int] = Query(None),
    user_id: Optional[int] = Query(None),
    from_date: Optional[date] = Query(None),
    to_date: Optional[date] = Query(None),
    search: Optional[str] = Query(None, description="설문 제목 검색"),
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=100),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """설문 응답 목록을 조회합니다."""
    query = (
        select(SurveyResponseModel)
        .options(
            selectinload(SurveyResponseModel.survey),
            selectinload(SurveyResponseModel.user),
            selectinload(SurveyResponseModel.answers),
        )
    )

    if survey_id:
        query = query.where(SurveyResponseModel.survey_id == survey_id)
    if user_id:
        query = query.where(SurveyResponseModel.user_id == user_id)
    if from_date:
        query = query.where(SurveyResponseModel.response_date >= from_date)
    if to_date:
        query = query.where(SurveyResponseModel.response_date <= to_date)

    if search:
        query = query.join(Survey).where(Survey.title.ilike(f"%{search}%"))

    query = query.order_by(SurveyResponseModel.submitted_at.desc())
    offset = (page - 1) * size
    query = query.offset(offset).limit(size)

    result = await db.execute(query)
    responses = result.scalars().all()

    items = []
    for r in responses:
        item = response_record_to_dict(r, include_answers=False)
        item["survey_title"] = r.survey.title if r.survey else None
        item["user_name"] = r.user.full_name if r.user else None
        item["answer_count"] = len(r.answers or [])
        items.append(item)

    return {"items": items, "page": page, "size": size, "total": len(items)}


@router.get("/responses/{response_id}", summary="설문 응답 상세 조회")
async def get_response(
    response_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """설문 응답 상세 정보를 조회합니다."""
    result = await db.execute(
        select(SurveyResponseModel)
        .options(
            selectinload(SurveyResponseModel.answers),
            selectinload(SurveyResponseModel.survey),
            selectinload(SurveyResponseModel.user),
        )
        .where(SurveyResponseModel.id == response_id)
    )
    response = result.scalar_one_or_none()
    if not response:
        raise HTTPException(status_code=404, detail="설문 응답을 찾을 수 없습니다.")

    result_dict = response_record_to_dict(response, include_answers=True)
    result_dict["survey_title"] = response.survey.title if response.survey else None
    result_dict["user_name"] = response.user.full_name if response.user else None
    return result_dict


@router.put("/responses/{response_id}", summary="설문 응답 수정")
async def update_response(
    response_id: int,
    data: SurveyResponseUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    설문 응답을 수정합니다.
    기존 답변을 모두 삭제하고 새 답변으로 교체합니다.
    연결된 출퇴근 기록이 있으면 날짜도 함께 업데이트합니다.
    """
    try:
        # 1. 응답 레코드 조회 (없으면 404)
        result = await db.execute(
            select(SurveyResponseModel)
            .options(
                selectinload(SurveyResponseModel.answers),
                selectinload(SurveyResponseModel.commute_record),
            )
            .where(SurveyResponseModel.id == response_id)
        )
        response = result.scalar_one_or_none()
        if not response:
            raise HTTPException(
                status_code=404,
                detail=f"설문 응답(id={response_id})을 찾을 수 없습니다."
            )

        # 2. response_date 업데이트 (제공된 경우)
        if data.response_date is not None:
            response.response_date = data.response_date

        response.updated_at = datetime.utcnow()

        # 3. 기존 답변 전체 삭제
        await db.execute(
            delete(SurveyAnswer).where(SurveyAnswer.response_id == response_id)
        )

        # 4. 새 답변 생성
        for ans_data in data.answers:
            answer = SurveyAnswer(
                response_id=response_id,
                question_id=ans_data.question_id,
                answer_text=ans_data.answer_text,
                answer_number=ans_data.answer_number,
                answer_json=ans_data.answer_json,
            )
            db.add(answer)

        # 5. 연결된 출퇴근 기록 업데이트 (날짜 변경 시)
        if data.response_date is not None and response.commute_record:
            commute = response.commute_record
            commute.record_date = data.response_date
            commute.record_year = data.response_date.year
            commute.record_month = data.response_date.month
            commute.updated_at = datetime.utcnow()

        await db.commit()

        # 6. 최신 상태로 다시 조회
        final_result = await db.execute(
            select(SurveyResponseModel)
            .options(
                selectinload(SurveyResponseModel.answers),
                selectinload(SurveyResponseModel.survey),
                selectinload(SurveyResponseModel.user),
            )
            .where(SurveyResponseModel.id == response_id)
        )
        updated_response = final_result.scalar_one()

        result_dict = response_record_to_dict(updated_response, include_answers=True)
        result_dict["survey_title"] = updated_response.survey.title if updated_response.survey else None
        result_dict["user_name"] = updated_response.user.full_name if updated_response.user else None
        return result_dict

    except HTTPException:
        raise
    except Exception as e:
        await db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"설문 응답 수정 중 오류가 발생했습니다: {str(e)}"
        )


@router.delete("/responses/{response_id}", summary="설문 응답 삭제")
async def delete_response(
    response_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """설문 응답을 삭제합니다."""
    result = await db.execute(
        select(SurveyResponseModel).where(SurveyResponseModel.id == response_id)
    )
    response = result.scalar_one_or_none()
    if not response:
        raise HTTPException(status_code=404, detail="설문 응답을 찾을 수 없습니다.")

    await db.delete(response)
    await db.commit()
    return {"success": True, "message": "설문 응답이 삭제되었습니다.", "response_id": response_id}


# ============================================================================
# Commute Records
# ============================================================================

@router.get("/commute/", summary="출퇴근 기록 목록 조회")
async def list_commute_records(
    year: Optional[int] = Query(None),
    month: Optional[int] = Query(None),
    user_id: Optional[int] = Query(None),
    department_id: Optional[int] = Query(None),
    page: int = Query(1, ge=1),
    size: int = Query(50, ge=1, le=200),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """출퇴근 기록 목록을 조회합니다."""
    query = (
        select(CommuteRecord)
        .options(
            selectinload(CommuteRecord.user).selectinload(User.department)
        )
    )

    if year:
        query = query.where(CommuteRecord.record_year == year)
    if month:
        query = query.where(CommuteRecord.record_month == month)
    if user_id:
        query = query.where(CommuteRecord.user_id == user_id)
    if department_id:
        query = query.join(User, CommuteRecord.user_id == User.id).where(
            User.department_id == department_id
        )

    query = query.order_by(CommuteRecord.record_date.desc(), CommuteRecord.user_id)
    offset = (page - 1) * size
    query = query.offset(offset).limit(size)

    result = await db.execute(query)
    records = result.scalars().all()

    return {
        "items": [commute_to_dict(r) for r in records],
        "page": page,
        "size": size,
        "total": len(records),
    }


@router.post("/commute/", summary="출퇴근 기록 생성")
async def create_commute_record(
    data: CommuteRecordCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """출퇴근 기록을 생성합니다."""
    user_id = data.user_id if data.user_id else current_user.id
    rec_date = data.record_date if data.record_date else date.today()

    transport = None
    if data.transport_method:
        try:
            transport = TransportMethod(data.transport_method)
        except ValueError:
            transport = None

    record = CommuteRecord(
        user_id=user_id,
        record_date=rec_date,
        record_year=rec_date.year,
        record_month=rec_date.month,
        check_in_time=data.check_in_time,
        check_out_time=data.check_out_time,
        departure_time=data.departure_time,
        arrival_time=data.arrival_time,
        transport_method=transport,
        commute_duration_minutes=data.commute_duration_minutes,
        note=data.note,
        survey_response_id=data.survey_response_id,
        is_late=data.is_late,
        is_early_leave=data.is_early_leave,
        is_absent=data.is_absent,
    )
    db.add(record)
    await db.commit()

    result = await db.execute(
        select(CommuteRecord)
        .options(selectinload(CommuteRecord.user).selectinload(User.department))
        .where(CommuteRecord.id == record.id)
    )
    record = result.scalar_one()
    return commute_to_dict(record)


@router.put("/commute/{record_id}", summary="출퇴근 기록 수정")
async def update_commute_record(
    record_id: int,
    data: CommuteRecordUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """출퇴근 기록을 수정합니다."""
    result = await db.execute(
        select(CommuteRecord)
        .options(selectinload(CommuteRecord.user).selectinload(User.department))
        .where(CommuteRecord.id == record_id)
    )
    record = result.scalar_one_or_none()
    if not record:
        raise HTTPException(status_code=404, detail="출퇴근 기록을 찾을 수 없습니다.")

    update_data = data.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        if field == "transport_method" and value is not None:
            try:
                value = TransportMethod(value)
            except ValueError:
                value = None
        setattr(record, field, value)

    record.updated_at = datetime.utcnow()
    await db.commit()

    result = await db.execute(
        select(CommuteRecord)
        .options(selectinload(CommuteRecord.user).selectinload(User.department))
        .where(CommuteRecord.id == record_id)
    )
    record = result.scalar_one()
    return commute_to_dict(record)


@router.delete("/commute/{record_id}", summary="출퇴근 기록 삭제")
async def delete_commute_record(
    record_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """출퇴근 기록을 삭제합니다."""
    result = await db.execute(
        select(CommuteRecord).where(CommuteRecord.id == record_id)
    )
    record = result.scalar_one_or_none()
    if not record:
        raise HTTPException(status_code=404, detail="출퇴근 기록을 찾을 수 없습니다.")

    await db.delete(record)
    await db.commit()
    return {"success": True, "message": "출퇴근 기록이 삭제되었습니다.", "record_id": record_id}


@router.post("/commute/check-in", summary="출근 체크인")
async def check_in(
    data: CheckInRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """현재 시각으로 출근 처리합니다. 오늘 기록이 있으면 체크인 시간을 업데이트합니다."""
    today = date.today()
    now_str = datetime.utcnow().strftime("%H:%M")

    # 오늘 기록 조회
    result = await db.execute(
        select(CommuteRecord).where(
            and_(
                CommuteRecord.user_id == current_user.id,
                CommuteRecord.record_date == today,
            )
        )
    )
    record = result.scalar_one_or_none()

    transport = None
    if data.transport_method:
        try:
            transport = TransportMethod(data.transport_method)
        except ValueError:
            pass

    if record:
        record.check_in_time = now_str
        if data.note:
            record.note = data.note
        if transport:
            record.transport_method = transport
        if data.departure_time:
            record.departure_time = data.departure_time
        record.updated_at = datetime.utcnow()
    else:
        record = CommuteRecord(
            user_id=current_user.id,
            record_date=today,
            record_year=today.year,
            record_month=today.month,
            check_in_time=now_str,
            departure_time=data.departure_time,
            transport_method=transport,
            note=data.note,
        )
        db.add(record)

    await db.commit()

    result = await db.execute(
        select(CommuteRecord)
        .options(selectinload(CommuteRecord.user).selectinload(User.department))
        .where(
            and_(
                CommuteRecord.user_id == current_user.id,
                CommuteRecord.record_date == today,
            )
        )
    )
    record = result.scalar_one()
    return commute_to_dict(record)


@router.post("/commute/check-out", summary="퇴근 체크아웃")
async def check_out(
    data: CheckOutRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """현재 시각으로 퇴근 처리합니다."""
    today = date.today()
    now_str = datetime.utcnow().strftime("%H:%M")

    result = await db.execute(
        select(CommuteRecord).where(
            and_(
                CommuteRecord.user_id == current_user.id,
                CommuteRecord.record_date == today,
            )
        )
    )
    record = result.scalar_one_or_none()

    if record:
        record.check_out_time = now_str
        if data.note:
            record.note = (record.note or "") + f" | 퇴근메모: {data.note}"
        record.updated_at = datetime.utcnow()
    else:
        record = CommuteRecord(
            user_id=current_user.id,
            record_date=today,
            record_year=today.year,
            record_month=today.month,
            check_out_time=now_str,
            note=data.note,
        )
        db.add(record)

    await db.commit()

    result = await db.execute(
        select(CommuteRecord)
        .options(selectinload(CommuteRecord.user).selectinload(User.department))
        .where(
            and_(
                CommuteRecord.user_id == current_user.id,
                CommuteRecord.record_date == today,
            )
        )
    )
    record = result.scalar_one()
    return commute_to_dict(record)


@router.get("/commute/summary", summary="월별 출퇴근 요약")
async def commute_summary(
    year: int = Query(..., ge=2000, le=2100),
    month: int = Query(..., ge=1, le=12),
    department_id: Optional[int] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """월별 직원별 출퇴근 요약을 반환합니다."""
    query = (
        select(
            CommuteRecord.user_id,
            func.count(CommuteRecord.id).label("total_days"),
            func.sum(
                cast(CommuteRecord.is_late, Integer)
            ).label("late_count"),
            func.sum(
                cast(CommuteRecord.is_early_leave, Integer)
            ).label("early_leave_count"),
            func.sum(
                cast(CommuteRecord.is_absent, Integer)
            ).label("absent_count"),
            func.avg(CommuteRecord.commute_duration_minutes).label("avg_commute"),
        )
        .where(
            and_(
                CommuteRecord.record_year == year,
                CommuteRecord.record_month == month,
            )
        )
        .group_by(CommuteRecord.user_id)
    )

    if department_id:
        query = query.join(User, CommuteRecord.user_id == User.id).where(
            User.department_id == department_id
        )

    result = await db.execute(query)
    rows = result.all()

    # Get user info
    user_ids = [r.user_id for r in rows]
    users_result = await db.execute(
        select(User)
        .options(selectinload(User.department))
        .where(User.id.in_(user_ids))
    )
    users_map = {u.id: u for u in users_result.scalars().all()}

    summary = []
    for row in rows:
        user = users_map.get(row.user_id)
        summary.append({
            "user_id": row.user_id,
            "user_name": user.full_name if user else "알 수 없음",
            "department_name": user.department.name if user and user.department else None,
            "year": year,
            "month": month,
            "total_days": row.total_days or 0,
            "late_count": int(row.late_count or 0),
            "early_leave_count": int(row.early_leave_count or 0),
            "absent_count": int(row.absent_count or 0),
            "avg_commute_minutes": float(row.avg_commute) if row.avg_commute else None,
        })

    return {"year": year, "month": month, "summary": summary}


@router.get("/commute/export/excel", summary="출퇴근 기록 Excel 내보내기")
async def export_commute_excel(
    year: int = Query(..., ge=2000, le=2100),
    month: int = Query(..., ge=1, le=12),
    department_id: Optional[int] = Query(None),
    user_id: Optional[int] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """출퇴근 기록을 Excel 파일로 내보냅니다."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="openpyxl 패키지가 설치되지 않았습니다. pip install openpyxl"
        )

    query = (
        select(CommuteRecord)
        .options(selectinload(CommuteRecord.user).selectinload(User.department))
        .where(
            and_(
                CommuteRecord.record_year == year,
                CommuteRecord.record_month == month,
            )
        )
    )
    if department_id:
        query = query.join(User, CommuteRecord.user_id == User.id).where(
            User.department_id == department_id
        )
    if user_id:
        query = query.where(CommuteRecord.user_id == user_id)

    query = query.order_by(CommuteRecord.record_date, CommuteRecord.user_id)
    result = await db.execute(query)
    records = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{year}년 {month}월 출퇴근"

    headers = [
        "날짜", "사원명", "부서", "출근시간", "퇴근시간",
        "출발시간", "도착시간", "이동수단", "통근시간(분)",
        "지각", "조퇴", "결근", "메모"
    ]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, rec in enumerate(records, 2):
        transport_str = (
            rec.transport_method.value
            if rec.transport_method and hasattr(rec.transport_method, "value")
            else str(rec.transport_method) if rec.transport_method else ""
        )
        ws.cell(row=row_idx, column=1, value=str(rec.record_date))
        ws.cell(row=row_idx, column=2, value=rec.user.full_name if rec.user else "")
        ws.cell(row=row_idx, column=3, value=rec.user.department.name if rec.user and rec.user.department else "")
        ws.cell(row=row_idx, column=4, value=rec.check_in_time or "")
        ws.cell(row=row_idx, column=5, value=rec.check_out_time or "")
        ws.cell(row=row_idx, column=6, value=rec.departure_time or "")
        ws.cell(row=row_idx, column=7, value=rec.arrival_time or "")
        ws.cell(row=row_idx, column=8, value=transport_str)
        ws.cell(row=row_idx, column=9, value=rec.commute_duration_minutes or "")
        ws.cell(row=row_idx, column=10, value="Y" if rec.is_late else "")
        ws.cell(row=row_idx, column=11, value="Y" if rec.is_early_leave else "")
        ws.cell(row=row_idx, column=12, value="Y" if rec.is_absent else "")
        ws.cell(row=row_idx, column=13, value=rec.note or "")

    # Auto column width
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 30)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = f"commute_{year}{month:02d}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ============================================================================
# Survey Templates
# ============================================================================

BUILT_IN_TEMPLATES = {
    "commute_daily": {
        "title": "일일 출퇴근 설문",
        "description": "매일 출퇴근 현황을 기록하는 설문입니다.",
        "category": "commute",
        "questions": [
            {"question_text": "오늘 출근 시간은 언제였나요?", "question_type": "time", "order": 0, "is_required": True},
            {"question_text": "오늘 퇴근 시간은 언제였나요?", "question_type": "time", "order": 1, "is_required": True},
            {"question_text": "출근 이동 수단은 무엇인가요?", "question_type": "single_choice", "order": 2, "is_required": True,
             "options": '["자가용","버스","지하철","자전거","도보","택시","기타"]'},
            {"question_text": "출근 소요 시간은 몇 분인가요?", "question_type": "number", "order": 3, "is_required": False,
             "placeholder": "예: 30"},
            {"question_text": "오늘 업무 만족도는 어떤가요?", "question_type": "scale", "order": 4, "is_required": False,
             "scale_min": 1, "scale_max": 5, "scale_min_label": "매우 불만족", "scale_max_label": "매우 만족"},
        ],
    },
    "satisfaction_monthly": {
        "title": "월간 직원 만족도 조사",
        "description": "매월 직원 만족도를 측정하는 설문입니다.",
        "category": "satisfaction",
        "questions": [
            {"question_text": "현재 업무에 대한 전반적인 만족도는?", "question_type": "scale", "order": 0, "is_required": True,
             "scale_min": 1, "scale_max": 10, "scale_min_label": "매우 불만족", "scale_max_label": "매우 만족"},
            {"question_text": "업무 환경에 만족하시나요?", "question_type": "scale", "order": 1, "is_required": True,
             "scale_min": 1, "scale_max": 5},
            {"question_text": "개선이 필요한 사항이 있다면 적어주세요.", "question_type": "textarea", "order": 2, "is_required": False,
             "placeholder": "자유롭게 작성해 주세요."},
        ],
    },
    "hr_onboarding": {
        "title": "신규 입사자 온보딩 설문",
        "description": "신규 입사자의 적응 상태를 파악하는 설문입니다.",
        "category": "hr",
        "questions": [
            {"question_text": "입사 후 첫 주 전반적인 경험은 어떠셨나요?", "question_type": "scale", "order": 0, "is_required": True,
             "scale_min": 1, "scale_max": 5},
            {"question_text": "업무 교육은 충분했나요?", "question_type": "single_choice", "order": 1, "is_required": True,
             "options": '["매우 충분","충분","보통","부족","매우 부족"]'},
            {"question_text": "동료 및 상사와의 관계는 어떤가요?", "question_type": "scale", "order": 2, "is_required": True,
             "scale_min": 1, "scale_max": 5},
            {"question_text": "추가적으로 필요한 지원이 있다면 적어주세요.", "question_type": "textarea", "order": 3, "is_required": False},
        ],
    },
    "general_feedback": {
        "title": "일반 피드백 설문",
        "description": "자유롭게 의견을 나눌 수 있는 설문입니다.",
        "category": "general",
        "questions": [
            {"question_text": "제목을 입력하세요.", "question_type": "text", "order": 0, "is_required": True},
            {"question_text": "내용을 입력하세요.", "question_type": "textarea", "order": 1, "is_required": True},
        ],
    },
    "custom_blank": {
        "title": "빈 설문 (직접 작성)",
        "description": "직접 문항을 추가하는 빈 설문 템플릿입니다.",
        "category": "custom",
        "questions": [],
    },
}


@router.get("/templates/", summary="설문 템플릿 목록")
async def list_templates(
    current_user: User = Depends(get_current_user),
):
    """내장 설문 템플릿 목록을 반환합니다."""
    items = []
    for template_type, tpl in BUILT_IN_TEMPLATES.items():
        items.append({
            "template_type": template_type,
            "title": tpl["title"],
            "description": tpl["description"],
            "category": tpl["category"],
            "question_count": len(tpl["questions"]),
        })
    return {"templates": items}


@router.post("/templates/{template_type}/create", summary="템플릿에서 설문 생성")
async def create_from_template(
    template_type: str,
    data: TemplateCreateRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """내장 템플릿을 기반으로 새 설문조사를 생성합니다."""
    if template_type not in BUILT_IN_TEMPLATES:
        raise HTTPException(
            status_code=404,
            detail=f"템플릿 '{template_type}'을(를) 찾을 수 없습니다. "
                   f"사용 가능한 템플릿: {list(BUILT_IN_TEMPLATES.keys())}"
        )

    tpl = BUILT_IN_TEMPLATES[template_type]
    title = data.title or tpl["title"]

    try:
        category_enum = SurveyCategory(tpl["category"])
    except ValueError:
        category_enum = SurveyCategory.GENERAL

    survey = Survey(
        title=title,
        description=tpl["description"],
        category=category_enum,
        status=SurveyStatus.DRAFT,
        target_department_id=data.target_department_id,
        is_anonymous=False,
        is_required=False,
        start_date=data.start_date,
        end_date=data.end_date,
        is_recurring=data.is_recurring,
        recurrence_type=data.recurrence_type,
        created_by=current_user.id,
    )
    db.add(survey)
    await db.flush()

    for q_data in tpl["questions"]:
        try:
            q_type = QuestionType(q_data.get("question_type", "text"))
        except ValueError:
            q_type = QuestionType.TEXT

        question = SurveyQuestion(
            survey_id=survey.id,
            question_text=q_data["question_text"],
            question_type=q_type,
            order=q_data.get("order", 0),
            is_required=q_data.get("is_required", False),
            options=q_data.get("options"),
            scale_min=q_data.get("scale_min"),
            scale_max=q_data.get("scale_max"),
            scale_min_label=q_data.get("scale_min_label"),
            scale_max_label=q_data.get("scale_max_label"),
            placeholder=q_data.get("placeholder"),
        )
        db.add(question)

    await db.commit()

    result = await db.execute(
        select(Survey)
        .options(selectinload(Survey.questions))
        .where(Survey.id == survey.id)
    )
    survey = result.scalar_one()
    return survey_to_response(survey)
