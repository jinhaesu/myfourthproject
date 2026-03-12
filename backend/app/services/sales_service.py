"""
Smart Finance Core - Sales Service
매출 자동화 & 전표 전환 서비스
"""
import io
import logging
from datetime import datetime, date
from decimal import Decimal
from typing import Optional, List, Tuple

import httpx
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_, func, or_
from sqlalchemy.orm import selectinload

from app.models.sales import (
    SalesChannel, SalesRecord, SalesAutomationSchedule,
    ChannelType, ApiType, SalesRecordStatus, ScheduleType
)
from app.models.accounting import Voucher, VoucherLine, VoucherStatus, TransactionType, Account
from app.core.config import settings

logger = logging.getLogger(__name__)


class SalesService:
    """
    매출 자동화 서비스
    - 판매 채널 CRUD
    - 매출 기록 관리 및 집계
    - 전표 전환
    - 엑셀 내보내기 / 이메일 발송
    - 자동화 스케줄 관리
    """

    def __init__(self, db: AsyncSession):
        self.db = db

    # ========================================================================
    # Channel CRUD
    # ========================================================================

    async def create_channel(
        self,
        code: str,
        name: str,
        channel_type: str,
        api_type: str,
        **kwargs
    ) -> SalesChannel:
        """판매 채널 생성"""
        # 중복 코드 확인
        existing = await self.db.execute(
            select(SalesChannel).where(SalesChannel.code == code)
        )
        if existing.scalar_one_or_none():
            raise ValueError(f"채널 코드 '{code}'가 이미 존재합니다.")

        channel = SalesChannel(
            code=code,
            name=name,
            channel_type=ChannelType(channel_type),
            api_type=ApiType(api_type),
            **kwargs
        )
        self.db.add(channel)
        await self.db.flush()
        return channel

    async def update_channel(
        self,
        channel_id: int,
        **updates
    ) -> SalesChannel:
        """판매 채널 수정"""
        channel = await self.db.get(SalesChannel, channel_id)
        if not channel:
            raise ValueError("판매 채널을 찾을 수 없습니다.")

        for field, value in updates.items():
            if value is not None:
                if field == "channel_type":
                    value = ChannelType(value)
                elif field == "api_type":
                    value = ApiType(value)
                setattr(channel, field, value)

        channel.updated_at = datetime.utcnow()
        await self.db.flush()
        return channel

    async def get_channels(self, active_only: bool = True) -> List[SalesChannel]:
        """판매 채널 목록 조회"""
        query = select(SalesChannel)
        if active_only:
            query = query.where(SalesChannel.is_active == True)
        query = query.order_by(SalesChannel.code)

        result = await self.db.execute(query)
        return list(result.scalars().all())

    async def get_channel(self, channel_id: int) -> Optional[SalesChannel]:
        """판매 채널 상세 조회"""
        return await self.db.get(SalesChannel, channel_id)

    async def delete_channel(self, channel_id: int) -> bool:
        """판매 채널 삭제 (소프트 삭제)"""
        channel = await self.db.get(SalesChannel, channel_id)
        if not channel:
            raise ValueError("판매 채널을 찾을 수 없습니다.")

        channel.is_active = False
        channel.updated_at = datetime.utcnow()
        await self.db.flush()
        return True

    # ========================================================================
    # Sales Records
    # ========================================================================

    async def upsert_sales_record(
        self,
        channel_id: int,
        year: int,
        month: int,
        data: dict
    ) -> SalesRecord:
        """매출 기록 upsert (unique key: channel_id, period_year, period_month)"""
        # 채널 존재 확인
        channel = await self.db.get(SalesChannel, channel_id)
        if not channel:
            raise ValueError("판매 채널을 찾을 수 없습니다.")

        # 기존 기록 조회
        result = await self.db.execute(
            select(SalesRecord).where(
                and_(
                    SalesRecord.channel_id == channel_id,
                    SalesRecord.period_year == year,
                    SalesRecord.period_month == month
                )
            )
        )
        record = result.scalar_one_or_none()

        if record:
            # 이미 전표 전환된 기록은 수정 불가
            if record.status == SalesRecordStatus.CONVERTED:
                raise ValueError("이미 전표 전환된 매출 기록은 수정할 수 없습니다.")

            # Update
            for field, value in data.items():
                if value is not None and hasattr(record, field):
                    setattr(record, field, value)
            record.updated_at = datetime.utcnow()
        else:
            # Insert
            record = SalesRecord(
                channel_id=channel_id,
                period_year=year,
                period_month=month,
                **data
            )
            self.db.add(record)

        await self.db.flush()
        return record

    async def get_sales_records(
        self,
        year: int,
        month: int,
        channel_id: Optional[int] = None,
        status: Optional[str] = None
    ) -> List[SalesRecord]:
        """매출 기록 조회"""
        query = select(SalesRecord).options(
            selectinload(SalesRecord.channel)
        ).where(
            and_(
                SalesRecord.period_year == year,
                SalesRecord.period_month == month
            )
        )

        if channel_id:
            query = query.where(SalesRecord.channel_id == channel_id)
        if status:
            query = query.where(SalesRecord.status == SalesRecordStatus(status))

        query = query.order_by(SalesRecord.channel_id)
        result = await self.db.execute(query)
        return list(result.scalars().all())

    async def confirm_sales_record(self, record_id: int) -> SalesRecord:
        """매출 기록 확정"""
        record = await self.db.get(SalesRecord, record_id)
        if not record:
            raise ValueError("매출 기록을 찾을 수 없습니다.")

        if record.status != SalesRecordStatus.PENDING:
            raise ValueError(f"대기(pending) 상태의 매출 기록만 확정할 수 있습니다. 현재 상태: {record.status.value}")

        record.status = SalesRecordStatus.CONFIRMED
        record.confirmed_at = datetime.utcnow()
        record.updated_at = datetime.utcnow()
        await self.db.flush()
        return record

    async def get_monthly_summary(self, year: int, month: int) -> dict:
        """전체 채널 월간 집계"""
        records = await self.get_sales_records(year, month)
        channels = await self.get_channels(active_only=False)
        channel_map = {ch.id: ch for ch in channels}

        channel_summaries = []
        total_gross = Decimal("0")
        total_returns = Decimal("0")
        total_net = Decimal("0")
        total_commission = Decimal("0")
        total_settlement = Decimal("0")
        total_orders = 0
        total_cancels = 0

        for record in records:
            ch = channel_map.get(record.channel_id)
            if not ch:
                continue

            channel_summaries.append({
                "channel_id": ch.id,
                "channel_code": ch.code,
                "channel_name": ch.name,
                "channel_type": ch.channel_type.value if hasattr(ch.channel_type, 'value') else str(ch.channel_type),
                "period_year": year,
                "period_month": month,
                "gross_sales": record.gross_sales,
                "returns": record.returns,
                "net_sales": record.net_sales,
                "commission": record.commission,
                "commission_rate": ch.commission_rate,
                "settlement_amount": record.settlement_amount,
                "order_count": record.order_count,
                "cancel_count": record.cancel_count,
                "status": record.status.value if hasattr(record.status, 'value') else str(record.status),
            })

            total_gross += record.gross_sales
            total_returns += record.returns
            total_net += record.net_sales
            total_commission += record.commission
            total_settlement += record.settlement_amount
            total_orders += record.order_count
            total_cancels += record.cancel_count

        return {
            "period_year": year,
            "period_month": month,
            "total_gross_sales": total_gross,
            "total_returns": total_returns,
            "total_net_sales": total_net,
            "total_commission": total_commission,
            "total_settlement": total_settlement,
            "total_orders": total_orders,
            "total_cancels": total_cancels,
            "channel_summaries": channel_summaries,
        }

    async def get_channel_trend(
        self,
        channel_id: int,
        months: int = 12
    ) -> dict:
        """채널별 추이"""
        channel = await self.db.get(SalesChannel, channel_id)
        if not channel:
            raise ValueError("판매 채널을 찾을 수 없습니다.")

        query = (
            select(SalesRecord)
            .where(SalesRecord.channel_id == channel_id)
            .order_by(SalesRecord.period_year.desc(), SalesRecord.period_month.desc())
            .limit(months)
        )

        result = await self.db.execute(query)
        records = list(result.scalars().all())

        # 시간순 정렬 (오래된 것 -> 최신)
        records.sort(key=lambda r: (r.period_year, r.period_month))

        trend = []
        for record in records:
            trend.append({
                "period_year": record.period_year,
                "period_month": record.period_month,
                "gross_sales": record.gross_sales,
                "net_sales": record.net_sales,
                "commission": record.commission,
                "settlement_amount": record.settlement_amount,
                "order_count": record.order_count,
            })

        return {
            "channel_id": channel.id,
            "channel_code": channel.code,
            "channel_name": channel.name,
            "trend": trend,
        }

    async def get_yearly_summary(self, year: int) -> dict:
        """연간 채널별 집계"""
        query = (
            select(
                SalesRecord.channel_id,
                func.sum(SalesRecord.gross_sales).label("total_gross_sales"),
                func.sum(SalesRecord.returns).label("total_returns"),
                func.sum(SalesRecord.net_sales).label("total_net_sales"),
                func.sum(SalesRecord.commission).label("total_commission"),
                func.sum(SalesRecord.settlement_amount).label("total_settlement"),
                func.sum(SalesRecord.order_count).label("total_orders"),
                func.sum(SalesRecord.cancel_count).label("total_cancels"),
            )
            .where(SalesRecord.period_year == year)
            .group_by(SalesRecord.channel_id)
        )

        result = await self.db.execute(query)
        rows = result.all()

        channels = await self.get_channels(active_only=False)
        channel_map = {ch.id: ch for ch in channels}

        channel_items = []
        grand_gross = Decimal("0")
        grand_net = Decimal("0")
        grand_commission = Decimal("0")
        grand_settlement = Decimal("0")

        for row in rows:
            ch = channel_map.get(row.channel_id)
            if not ch:
                continue

            total_gross = row.total_gross_sales or Decimal("0")
            total_returns = row.total_returns or Decimal("0")
            total_net = row.total_net_sales or Decimal("0")
            total_commission = row.total_commission or Decimal("0")
            total_settlement = row.total_settlement or Decimal("0")
            total_orders = row.total_orders or 0
            total_cancels = row.total_cancels or 0

            channel_items.append({
                "channel_id": ch.id,
                "channel_code": ch.code,
                "channel_name": ch.name,
                "total_gross_sales": total_gross,
                "total_returns": total_returns,
                "total_net_sales": total_net,
                "total_commission": total_commission,
                "total_settlement": total_settlement,
                "total_orders": total_orders,
                "total_cancels": total_cancels,
            })

            grand_gross += total_gross
            grand_net += total_net
            grand_commission += total_commission
            grand_settlement += total_settlement

        return {
            "year": year,
            "grand_total_gross_sales": grand_gross,
            "grand_total_net_sales": grand_net,
            "grand_total_commission": grand_commission,
            "grand_total_settlement": grand_settlement,
            "channels": channel_items,
        }

    # ========================================================================
    # Voucher Conversion (전표 전환)
    # ========================================================================

    async def convert_to_voucher(
        self,
        record_ids: List[int],
        user_id: int,
        department_id: int,
        description: Optional[str] = None
    ) -> dict:
        """매출 기록을 전표로 전환"""
        voucher_ids = []
        converted_count = 0

        for record_id in record_ids:
            record = await self.db.execute(
                select(SalesRecord).options(
                    selectinload(SalesRecord.channel)
                ).where(SalesRecord.id == record_id)
            )
            record = record.scalar_one_or_none()

            if not record:
                raise ValueError(f"매출 기록(ID: {record_id})을 찾을 수 없습니다.")

            if record.status == SalesRecordStatus.CONVERTED:
                raise ValueError(
                    f"매출 기록(ID: {record_id})은 이미 전표 전환되었습니다."
                )

            if record.status not in (SalesRecordStatus.CONFIRMED, SalesRecordStatus.SETTLED):
                raise ValueError(
                    f"매출 기록(ID: {record_id})은 확정(confirmed) 또는 정산(settled) 상태여야 전표 전환할 수 있습니다."
                )

            channel = record.channel
            voucher_date = date(record.period_year, record.period_month, 1)

            # 매출채권 계정 조회 (120100)
            ar_account = await self._get_account_by_code("120100")
            # 상품매출 계정 조회 (410100)
            sales_account = await self._get_account_by_code("410100")
            # 판매수수료 계정 조회 (813700 - 지급수수료)
            commission_account = await self._get_account_by_code("813700")

            voucher_desc = description or f"{channel.name} {record.period_year}년 {record.period_month}월 매출"

            # 전표번호 생성
            voucher_number = await self._generate_voucher_number(voucher_date)

            # --- 매출 전표 ---
            # 차변: 매출채권 (순매출)
            # 대변: 상품매출 (순매출)
            voucher = Voucher(
                voucher_number=voucher_number,
                voucher_date=voucher_date,
                transaction_date=voucher_date,
                description=voucher_desc,
                transaction_type=TransactionType.GENERAL,
                department_id=department_id,
                created_by=user_id,
                total_debit=record.net_sales,
                total_credit=record.net_sales,
                status=VoucherStatus.DRAFT,
            )
            self.db.add(voucher)
            await self.db.flush()

            # 차변: 매출채권
            debit_line = VoucherLine(
                voucher_id=voucher.id,
                line_number=1,
                account_id=ar_account.id,
                debit_amount=record.net_sales,
                credit_amount=Decimal("0"),
                description=f"{channel.name} 매출채권",
            )
            self.db.add(debit_line)

            # 대변: 상품매출
            credit_line = VoucherLine(
                voucher_id=voucher.id,
                line_number=2,
                account_id=sales_account.id,
                debit_amount=Decimal("0"),
                credit_amount=record.net_sales,
                description=f"{channel.name} 상품매출",
            )
            self.db.add(credit_line)

            voucher_ids.append(voucher.id)

            # --- 수수료 전표 (수수료가 있는 경우) ---
            if record.commission > Decimal("0"):
                commission_voucher_number = await self._generate_voucher_number(voucher_date)
                commission_desc = f"{channel.name} {record.period_year}년 {record.period_month}월 판매수수료"

                commission_voucher = Voucher(
                    voucher_number=commission_voucher_number,
                    voucher_date=voucher_date,
                    transaction_date=voucher_date,
                    description=commission_desc,
                    transaction_type=TransactionType.GENERAL,
                    department_id=department_id,
                    created_by=user_id,
                    total_debit=record.commission,
                    total_credit=record.commission,
                    status=VoucherStatus.DRAFT,
                )
                self.db.add(commission_voucher)
                await self.db.flush()

                # 차변: 판매수수료
                comm_debit = VoucherLine(
                    voucher_id=commission_voucher.id,
                    line_number=1,
                    account_id=commission_account.id,
                    debit_amount=record.commission,
                    credit_amount=Decimal("0"),
                    description=f"{channel.name} 판매수수료",
                )
                self.db.add(comm_debit)

                # 대변: 매출채권
                comm_credit = VoucherLine(
                    voucher_id=commission_voucher.id,
                    line_number=2,
                    account_id=ar_account.id,
                    debit_amount=Decimal("0"),
                    credit_amount=record.commission,
                    description=f"{channel.name} 수수료 상계",
                )
                self.db.add(comm_credit)

                voucher_ids.append(commission_voucher.id)

            # 매출 기록 상태 업데이트
            record.status = SalesRecordStatus.CONVERTED
            record.converted_at = datetime.utcnow()
            record.voucher_id = voucher.id
            record.updated_at = datetime.utcnow()

            converted_count += 1

        await self.db.flush()

        return {
            "converted_count": converted_count,
            "voucher_ids": voucher_ids,
            "message": f"{converted_count}건의 매출 기록이 전표로 전환되었습니다.",
        }

    async def _get_account_by_code(self, code: str) -> Account:
        """계정과목 코드로 조회"""
        result = await self.db.execute(
            select(Account).where(
                Account.code == code,
                Account.is_active == True
            )
        )
        account = result.scalar_one_or_none()
        if not account:
            raise ValueError(f"계정과목({code})이 설정되지 않았습니다. 계정과목을 확인하세요.")
        return account

    async def _generate_voucher_number(self, voucher_date: date) -> str:
        """전표번호 생성"""
        prefix = f"V{voucher_date.strftime('%Y%m%d')}"

        result = await self.db.execute(
            select(Voucher).where(
                Voucher.voucher_number.like(f"{prefix}%")
            ).order_by(Voucher.voucher_number.desc())
            .limit(1)
        )
        last_voucher = result.scalars().first()

        if last_voucher:
            last_seq = int(last_voucher.voucher_number[-4:])
            new_seq = last_seq + 1
        else:
            new_seq = 1

        return f"{prefix}{new_seq:04d}"

    # ========================================================================
    # Excel Export
    # ========================================================================

    async def export_monthly_excel(self, year: int, month: int) -> bytes:
        """월간 매출 엑셀 내보내기"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            raise ValueError("openpyxl 패키지가 설치되어 있지 않습니다. pip install openpyxl")

        summary = await self.get_monthly_summary(year, month)
        wb = Workbook()

        # ---- Sheet 1: 채널별 매출 요약 ----
        ws1 = wb.active
        ws1.title = "매출 요약"

        # 헤더 스타일
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # 제목
        ws1.merge_cells("A1:I1")
        title_cell = ws1["A1"]
        title_cell.value = f"{year}년 {month}월 채널별 매출 요약"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")

        # 헤더
        headers = [
            "채널", "총매출", "반품/환불", "순매출",
            "수수료율(%)", "수수료", "정산금액", "주문건수", "취소건수"
        ]
        for col, header in enumerate(headers, 1):
            cell = ws1.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 데이터
        row = 4
        for ch_summary in summary["channel_summaries"]:
            ws1.cell(row=row, column=1, value=ch_summary["channel_name"]).border = thin_border
            ws1.cell(row=row, column=2, value=float(ch_summary["gross_sales"])).border = thin_border
            ws1.cell(row=row, column=3, value=float(ch_summary["returns"])).border = thin_border
            ws1.cell(row=row, column=4, value=float(ch_summary["net_sales"])).border = thin_border
            ws1.cell(row=row, column=5, value=float(ch_summary["commission_rate"])).border = thin_border
            ws1.cell(row=row, column=6, value=float(ch_summary["commission"])).border = thin_border
            ws1.cell(row=row, column=7, value=float(ch_summary["settlement_amount"])).border = thin_border
            ws1.cell(row=row, column=8, value=ch_summary["order_count"]).border = thin_border
            ws1.cell(row=row, column=9, value=ch_summary["cancel_count"]).border = thin_border

            # 금액 포맷
            for col in [2, 3, 4, 6, 7]:
                ws1.cell(row=row, column=col).number_format = '#,##0'
            ws1.cell(row=row, column=5).number_format = '0.0'

            row += 1

        # 합계 행
        total_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
        total_font = Font(bold=True)
        ws1.cell(row=row, column=1, value="합계").font = total_font
        ws1.cell(row=row, column=1).fill = total_fill
        ws1.cell(row=row, column=2, value=float(summary["total_gross_sales"]))
        ws1.cell(row=row, column=3, value=float(summary["total_returns"]))
        ws1.cell(row=row, column=4, value=float(summary["total_net_sales"]))
        ws1.cell(row=row, column=5, value="")
        ws1.cell(row=row, column=6, value=float(summary["total_commission"]))
        ws1.cell(row=row, column=7, value=float(summary["total_settlement"]))
        ws1.cell(row=row, column=8, value=summary["total_orders"])
        ws1.cell(row=row, column=9, value=summary["total_cancels"])

        for col in range(1, 10):
            ws1.cell(row=row, column=col).font = total_font
            ws1.cell(row=row, column=col).fill = total_fill
            ws1.cell(row=row, column=col).border = thin_border
        for col in [2, 3, 4, 6, 7]:
            ws1.cell(row=row, column=col).number_format = '#,##0'

        # 열 너비 조정
        ws1.column_dimensions['A'].width = 20
        for col_letter in ['B', 'C', 'D', 'E', 'F', 'G']:
            ws1.column_dimensions[col_letter].width = 15
        ws1.column_dimensions['H'].width = 12
        ws1.column_dimensions['I'].width = 12

        # ---- Sheet 2: 채널별 상세 데이터 ----
        ws2 = wb.create_sheet("채널별 상세")

        ws2.merge_cells("A1:H1")
        title_cell2 = ws2["A1"]
        title_cell2.value = f"{year}년 {month}월 채널별 상세 데이터"
        title_cell2.font = Font(bold=True, size=14)
        title_cell2.alignment = Alignment(horizontal="center")

        detail_headers = [
            "채널코드", "채널명", "유형", "총매출",
            "순매출", "수수료", "정산금액", "상태"
        ]
        for col, header in enumerate(detail_headers, 1):
            cell = ws2.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        row2 = 4
        for ch_summary in summary["channel_summaries"]:
            ws2.cell(row=row2, column=1, value=ch_summary.get("channel_code", "")).border = thin_border
            ws2.cell(row=row2, column=2, value=ch_summary["channel_name"]).border = thin_border
            ws2.cell(row=row2, column=3, value=ch_summary.get("channel_type", "")).border = thin_border
            ws2.cell(row=row2, column=4, value=float(ch_summary["gross_sales"])).border = thin_border
            ws2.cell(row=row2, column=5, value=float(ch_summary["net_sales"])).border = thin_border
            ws2.cell(row=row2, column=6, value=float(ch_summary["commission"])).border = thin_border
            ws2.cell(row=row2, column=7, value=float(ch_summary["settlement_amount"])).border = thin_border
            ws2.cell(row=row2, column=8, value=ch_summary["status"]).border = thin_border

            for col in [4, 5, 6, 7]:
                ws2.cell(row=row2, column=col).number_format = '#,##0'

            row2 += 1

        # 열 너비 조정
        for col_letter in ['A', 'B', 'C']:
            ws2.column_dimensions[col_letter].width = 18
        for col_letter in ['D', 'E', 'F', 'G']:
            ws2.column_dimensions[col_letter].width = 15
        ws2.column_dimensions['H'].width = 12

        # 바이트로 저장
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    # ========================================================================
    # Email Report
    # ========================================================================

    async def send_monthly_report(
        self,
        year: int,
        month: int,
        recipients: List[str],
        subject: Optional[str] = None
    ) -> bool:
        """월간 매출 리포트 이메일 발송"""
        # 엑셀 생성
        excel_bytes = await self.export_monthly_excel(year, month)

        email_subject = subject or f"[Smart Finance] {year}년 {month}월 매출 리포트"

        # 요약 데이터
        summary = await self.get_monthly_summary(year, month)

        html_body = f"""
        <div style="font-family: 'Apple SD Gothic Neo', sans-serif; max-width: 600px; margin: 0 auto; padding: 20px;">
            <div style="background: linear-gradient(135deg, #2563eb, #4f46e5); border-radius: 12px; padding: 24px; text-align: center; color: white;">
                <h1 style="margin: 0 0 8px; font-size: 20px;">Smart Finance Core</h1>
                <p style="margin: 0; opacity: 0.9; font-size: 14px;">{year}년 {month}월 매출 리포트</p>
            </div>
            <div style="background: #ffffff; border: 1px solid #e5e7eb; border-radius: 12px; margin-top: 16px; padding: 24px;">
                <h2 style="font-size: 16px; color: #1f2937;">매출 요약</h2>
                <table style="width: 100%; border-collapse: collapse; margin-top: 12px;">
                    <tr style="background: #f9fafb;">
                        <td style="padding: 8px; border: 1px solid #e5e7eb;">총매출</td>
                        <td style="padding: 8px; border: 1px solid #e5e7eb; text-align: right;">{summary['total_gross_sales']:,.0f}원</td>
                    </tr>
                    <tr>
                        <td style="padding: 8px; border: 1px solid #e5e7eb;">순매출</td>
                        <td style="padding: 8px; border: 1px solid #e5e7eb; text-align: right;">{summary['total_net_sales']:,.0f}원</td>
                    </tr>
                    <tr style="background: #f9fafb;">
                        <td style="padding: 8px; border: 1px solid #e5e7eb;">총수수료</td>
                        <td style="padding: 8px; border: 1px solid #e5e7eb; text-align: right;">{summary['total_commission']:,.0f}원</td>
                    </tr>
                    <tr>
                        <td style="padding: 8px; border: 1px solid #e5e7eb;">정산금액</td>
                        <td style="padding: 8px; border: 1px solid #e5e7eb; text-align: right;">{summary['total_settlement']:,.0f}원</td>
                    </tr>
                    <tr style="background: #f9fafb;">
                        <td style="padding: 8px; border: 1px solid #e5e7eb;">주문건수</td>
                        <td style="padding: 8px; border: 1px solid #e5e7eb; text-align: right;">{summary['total_orders']:,}건</td>
                    </tr>
                </table>
                <p style="color: #9ca3af; font-size: 12px; margin: 16px 0 0;">
                    상세 데이터는 첨부된 엑셀 파일을 확인하세요.
                </p>
            </div>
        </div>
        """

        if not settings.RESEND_API_KEY:
            logger.warning("RESEND_API_KEY not configured. Report email not sent.")
            logger.info(f"[DEBUG] Would send report to: {recipients}")
            return True

        import base64
        try:
            attachment_b64 = base64.b64encode(excel_bytes).decode("utf-8")

            async with httpx.AsyncClient() as client:
                response = await client.post(
                    "https://api.resend.com/emails",
                    headers={
                        "Authorization": f"Bearer {settings.RESEND_API_KEY}",
                        "Content-Type": "application/json",
                    },
                    json={
                        "from": settings.RESEND_FROM_EMAIL,
                        "to": recipients,
                        "subject": email_subject,
                        "html": html_body,
                        "attachments": [
                            {
                                "filename": f"sales_report_{year}_{month:02d}.xlsx",
                                "content": attachment_b64,
                            }
                        ],
                    },
                    timeout=30.0,
                )

            if response.status_code in (200, 201):
                logger.info(f"Sales report email sent to {recipients}")
                return True
            else:
                logger.error(f"Resend API error: {response.status_code} {response.text}")
                return False

        except Exception as e:
            logger.error(f"Failed to send report email: {e}")
            return False

    # ========================================================================
    # Excel Import
    # ========================================================================

    async def import_from_excel(self, file_bytes: bytes) -> dict:
        """엑셀 파일에서 매출 데이터 일괄 등록"""
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ValueError("openpyxl 패키지가 설치되어 있지 않습니다. pip install openpyxl")

        wb = load_workbook(io.BytesIO(file_bytes))
        ws = wb.active

        created_count = 0
        updated_count = 0
        error_count = 0
        errors = []

        # 헤더 확인 (2행부터 데이터)
        # 예상 컬럼: 채널코드, 연도, 월, 총매출, 반품, 순매출, 수수료, 정산금액, 주문건수, 취소건수
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[0]:
                continue

            try:
                channel_code = str(row[0]).strip()
                year = int(row[1])
                month = int(row[2])

                # 채널 조회
                result = await self.db.execute(
                    select(SalesChannel).where(SalesChannel.code == channel_code)
                )
                channel = result.scalar_one_or_none()
                if not channel:
                    raise ValueError(f"채널 코드 '{channel_code}'를 찾을 수 없습니다.")

                data = {
                    "gross_sales": Decimal(str(row[3] or 0)),
                    "returns": Decimal(str(row[4] or 0)),
                    "net_sales": Decimal(str(row[5] or 0)),
                    "commission": Decimal(str(row[6] or 0)),
                    "settlement_amount": Decimal(str(row[7] or 0)),
                    "order_count": int(row[8] or 0),
                    "cancel_count": int(row[9] or 0),
                }

                # upsert 확인
                existing = await self.db.execute(
                    select(SalesRecord).where(
                        and_(
                            SalesRecord.channel_id == channel.id,
                            SalesRecord.period_year == year,
                            SalesRecord.period_month == month
                        )
                    )
                )
                is_update = existing.scalar_one_or_none() is not None

                await self.upsert_sales_record(channel.id, year, month, data)

                if is_update:
                    updated_count += 1
                else:
                    created_count += 1

            except Exception as e:
                error_count += 1
                errors.append({
                    "row": row_idx,
                    "error": str(e)
                })

        await self.db.flush()

        return {
            "created_count": created_count,
            "updated_count": updated_count,
            "error_count": error_count,
            "errors": errors,
        }

    # ========================================================================
    # Automation Schedules
    # ========================================================================

    async def create_schedule(
        self,
        name: str,
        schedule_type: str,
        **kwargs
    ) -> SalesAutomationSchedule:
        """자동화 스케줄 생성"""
        schedule = SalesAutomationSchedule(
            name=name,
            schedule_type=ScheduleType(schedule_type),
            **kwargs
        )
        self.db.add(schedule)
        await self.db.flush()
        return schedule

    async def update_schedule(
        self,
        schedule_id: int,
        **updates
    ) -> SalesAutomationSchedule:
        """자동화 스케줄 수정"""
        schedule = await self.db.get(SalesAutomationSchedule, schedule_id)
        if not schedule:
            raise ValueError("스케줄을 찾을 수 없습니다.")

        for field, value in updates.items():
            if value is not None:
                if field == "schedule_type":
                    value = ScheduleType(value)
                setattr(schedule, field, value)

        schedule.updated_at = datetime.utcnow()
        await self.db.flush()
        return schedule

    async def get_schedules(self) -> List[SalesAutomationSchedule]:
        """자동화 스케줄 목록 조회"""
        result = await self.db.execute(
            select(SalesAutomationSchedule).order_by(SalesAutomationSchedule.id)
        )
        return list(result.scalars().all())

    async def delete_schedule(self, schedule_id: int) -> bool:
        """자동화 스케줄 삭제"""
        schedule = await self.db.get(SalesAutomationSchedule, schedule_id)
        if not schedule:
            raise ValueError("스케줄을 찾을 수 없습니다.")

        await self.db.delete(schedule)
        await self.db.flush()
        return True
